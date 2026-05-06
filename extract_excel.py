import openpyxl

def extract_excel_info():
    print("Abriendo Excel...")
    try:
        wb = openpyxl.load_workbook('series (3).xlsm', read_only=True, data_only=True)
        with open("excel_info.txt", "w", encoding="utf-8") as f:
            f.write(f"Hojas: {wb.sheetnames}\n\n")
            for sheet_name in wb.sheetnames:
                f.write(f"=== Hoja: '{sheet_name}' ===\n")
                ws = wb[sheet_name]
                row_count = 0
                for row in ws.iter_rows(max_row=20, values_only=True):
                    if any(v is not None for v in row):
                        f.write(f"  Fila {row_count}: {row[:10]}\n")
                    row_count += 1
        print("Excel info extract successfully")
    except Exception as e:
        with open("excel_error.txt", "w", encoding="utf-8") as f:
            f.write(str(e))
        print("Error reading excel")

extract_excel_info()
