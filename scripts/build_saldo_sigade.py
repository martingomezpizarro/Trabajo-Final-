"""
Serie historica de saldo de deuda publica argentina (SIGADE .mdb)
Hojas de salida:
  A.1 - Deuda bruta por instrumento y situacion de pago
  A.3 - Deuda en situacion de pago normal por moneda y tasa
Unidades: millones de USD
Frecuencia: trimestral (un archivo .mdb por periodo)
"""

import re
import warnings
from pathlib import Path
from datetime import date

import pandas as pd
import pyodbc

warnings.filterwarnings("ignore", category=UserWarning)

MDB_DIR     = Path(r"c:\Users\Usuario\Desktop\MARTIN\ECONOMICS\TRABAJO FINAL\Trabajo\data\basesingade deuda")
OUTPUT_PATH = Path(r"c:\Users\Usuario\Desktop\MARTIN\ECONOMICS\TRABAJO FINAL\Trabajo\data\Variables Finales\saldo_deuda.xlsx")

# ============================================================
# Column aliases
# ============================================================

CLASIF_ALIASES = [
    "Clasificacion de la Deuda", "Clasificaci\xf3n de la Deuda",
    "Clasificación de la Deuda",
    "Tipo de Deuda", "Tipo de deuda",
    "Etiquetas de fila",             # pivot-table variant (12/2015, 03/2016)
    "Boletin", "BOLETIN",            # pivot-table variant (06/2016, 09/2016)
    "Bolet\xedn", "BOLET\xcdN",     # same with accented í/Í
]
OPERACION_ALIASES = [
    "Nombre de la Operacion", "Nombre de la Operaci\xf3n",
    "Nombre de la Operación", "Nombre de la Operacio\xf3n",
]
TASA_ALIASES = [
    "Tipo de Tasa", "Tipo de TAsa", "TAsa", "Tasa",
]
MONEDA_ALIASES = [
    "Moneda del Prestamo", "Moneda del prestamo", "Moneda de Origen",
    "Moneda del Tramo",
]
SALDO_USD_ALIASES = [
    "Saldo en Dolares", "Saldo en D\xf3lares", "Saldo en dolares",
    "saldo en dolares", "Saldo  en D\xf3lares",   # double-space variant (2011)
    "Saldo  en Dolares",
    "dolares", "Monto en dolares", "Monto en d\xf3lares",
    "SumaDesaldo en dolares",                      # 2009-2010 variant
    "Suma de Monto en dolares", "Suma de Monto en d\xf3lares",   # 2015-2016 pivot tables
    "Suma de Montoen D\xf3lares", "Suma de Montoen Dolares",     # same, space-elided form
    "Suma de Suma de Monto en M de Origen",                       # fallback if USD col absent
]

# ============================================================
# Category normalization (2007-2011 Tipo de Deuda -> standard)
# ============================================================

CAT_MAP = {
    "ANTICIPOS DEL BANCO CENTRAL - Cto.Plazo": "ANTICIPOS BCRA",
    "BANCA COMERCIAL -BANCA PRIVADA EXTERNA":   "BANCA",
    "BANCA COMERCIAL-BANCA PRIVADA INTERNA":    "BANCA",
    "BANCA COMERCIAL- BANCA PRIVADA INTERNA":   "BANCA",
    "BONOS ESPA\xd1OLES CON GARANTIA.":         "BONOS",
    "BONOS ESPANOLES CON GARANTIA.":            "BONOS",
    "DECRETO 977 - BOGAR 2020":                 "DECRETO 977",
    "LETRAS DEL TESORO":                        "LETRAS DEL TESORO",
    "ORGANISM. OFICIALES - BILATERAL INTERNA":  "BILATERALES",
    "ORGANISM. OFICIALES -BILATERAL EXTERNA":   "BILATERALES",
    "ORGANISM. OFICIALES -CLUB DE PARIS":       "CLUB DE PARIS",
    "ORGANISMOS INTERNACIONALES - BEI":         "BEI",
    "ORGANISMOS INTERNACIONALES - BID/FIDA":    "BID",
    "ORGANISMOS INTERNACIONALES - CAF":         "CAF",
    "ORGANISMOS INTERNACIONALES - FIDA":        "FIDA",
    "ORGANISMOS INTERNACIONALES -BIRF":         "BIRF",
    "ORGANISMOS INTERNACIONALES -FONPLATA":     "FONPLATA",
    "OTROS ACREEDORES-PROVEEDORES EXTERNA":     "OTROS",
    "OTROS ACREEDORES-PROVEEDORES INTERNA":     "OTROS",
    "OTROS ACREEDORES-Provincias Dto. 1023":    "OTROS",
    "PAGARES":                                  "PAGARES",
    "PRESTAMO GARANTIZADO TASA FIJA":           "PRESTAMOS GARANTIZADOS",
    "PRESTAMO GARANTIZADO TASA FIJA VTO 2011":  "PRESTAMOS GARANTIZADOS",
    "PRESTAMO GARANTIZADO TASA VARIABLE":       "PRESTAMOS GARANTIZADOS",
    "TITULOS PUBLICOS -Bonos LP":               "BONOS",
    "TITULOS PUBLICOS -Bonos Pesificados":       "BONOS",
    "LETES NO ELEGIBLES":                       "BONOS",
    "CCF NO ELEGIBLE":                          "BONOS",
    "BANCA":                                    "BANCA",
    "OTROS":                                    "OTROS",
    "BONOS":                                    "BONOS",
}

# ============================================================
# Subcategory patterns per Clasificacion
# ============================================================

SUBCAT_PATTERNS = {
    "BONOS": [
        (r"^BOCON",                                    "BOCON"),
        (r"^BODEN",                                    "BODEN"),
        (r"^BONAR",                                    "BONAR"),
        (r"^BONTE",                                    "BONTES"),
        (r"^BONO CONSOLIDADO",                         "BONO CONSOLIDADO"),
        (r"^BONO DEL TESORO",                          "BONO DEL TESORO"),
        (r"^BONO R\.A\.",                              "BONO R.A."),
        (r"^(GLOBAL BOND|BONOS GLOBALES|GLOBAL B\.|GLOBAL/)", "GLOBAL BOND"),
        (r"^CUASIPAR",                                 "CUASIPAR"),
        (r"^DISCOUNT",                                 "DISCOUNT"),
        (r"^LETRA INTRANSFERIBLE",                     "LETRA INTRANSFERIBLE"),
        (r"^LETRA/U\$S/F\.DES",                       "LETRAS U$S F.DESEND."),
        (r"^LETRA/U\$S/ORG\.MULT",                    "LETRAS U$S ORG.MULT."),
        (r"^OCMO",                                     "OCMO"),
        (r"^PAR/",                                     "PAR"),
        (r"^(PRE|PRO)\d",                              "BONOS REESTRUCTURADOS"),
        (r"^BIRAD",                                    "BIRAD"),
        (r"^BIRA[EF]",                                 "BIRA"),
    ],
    "LETRAS DEL TESORO": [
        (r"^LETRA/\$",   "Moneda nacional"),
        (r"^LETRA/U\$S", "Moneda extranjera"),
        (r"^LECAP",      "LECAP"),
        (r"^LETES",      "LETES"),
        (r"^LETRA FFRH", "LETRA FFRH"),
        (r"^LETRA FFSIT","LETRA FFSIT"),
        (r"^LETRA ANSES","LETRA ANSES"),
        (r"^LETRA FGS",  "LETRA FGS"),
        (r"^LETRA",      "Otras Letras"),
    ],
    "BANCA": [
        (r"FINANC\.?\s*BNA|FINANC\. BNA",  "Financiaciones BNA"),
        (r"^PAGARE",                         "Pagares"),
    ],
    "PRESTAMOS GARANTIZADOS": [
        (r"T\.FIJA",    "Tasa Fija"),
        (r"T\.VARIABLE","Tasa Variable"),
    ],
}

# ============================================================
# Moneda and Tasa groupings for A.3
# ============================================================

def moneda_group(m: str) -> tuple[str, str]:
    """Returns (level1, level2) for A.3 structure.

    Handles both ISO codes (all periods) and Spanish full names that appear
    in 2008 Q2-Q4 and later 2007-era SALDOS tables.
    """
    import unicodedata
    m = str(m).strip().upper()
    # Strip accents for robust comparison (e.g. DÓLAR -> DOLAR)
    m = "".join(c for c in unicodedata.normalize("NFD", m)
                if unicodedata.category(c) != "Mn")

    if m in ("ARP", "PESO ARGENTINO"):
        return ("Moneda local", "ARP (no ajustable CER)")
    if m in ("UCP", "PESO ARGENTINO + CER"):
        return ("Moneda local", "UCP (ajustable CER)")
    if m in ("USD", "DOLARES ESTADOUNIDENSES"):
        return ("Moneda extranjera", "Dolares")
    if m in ("EUR", "EURO"):
        return ("Moneda extranjera", "Euros")
    if m in ("JPY", "YENES JAPONESES"):
        return ("Moneda extranjera", "Yenes")
    if m in ("SDR", "DEG", "DERECHOS ESPECIALES DE GIRO",
             "BID", "UNIDAD DE CUENTA DEL B.I.D."):
        return ("Moneda extranjera", "DEG")
    return ("Moneda extranjera", "Otras monedas")


def tasa_group(t: str) -> str:
    t = str(t).strip().upper()
    if "FIJA" in t:
        return "Tasa Fija"
    if "CERO" in t:
        return "Tasa Cero"
    return "Tasa Variable"

# ============================================================
# Helpers
# ============================================================

def parse_file_date(stem: str) -> date | None:
    m = re.search(r"(\d{4})-(\d{2})-(\d{2})$", stem)
    if not m:
        return None
    y, a, b = int(m.group(1)), int(m.group(2)), int(m.group(3))
    day, month = (a, b) if a > 12 else (b, a)
    try:
        return date(y, month, day)
    except ValueError:
        return None


def col_to_date(col: str) -> date:
    mm, yyyy = col.split("/")
    return date(int(yyyy), int(mm), 1)


def pick_col(cols: list[str], aliases: list[str]) -> str | None:
    col_map = {c.lower().strip(): c for c in cols}
    for a in aliases:
        if a in cols:
            return a
        if a.lower().strip() in col_map:
            return col_map[a.lower().strip()]
    return None


def find_table(tables: list[str], must: list[str], exclude: list[str] = None) -> str | None:
    exclude = exclude or []
    for t in tables:
        tu = t.upper()
        if any(k.upper() in tu for k in exclude):
            continue
        if all(k.upper() in tu for k in must):
            return t
    return None


def get_subcategory(clasif: str, operacion: str) -> str | None:
    patterns = SUBCAT_PATTERNS.get(clasif, [])
    for pat, subcat in patterns:
        if re.search(pat, str(operacion), re.IGNORECASE):
            return subcat
    return None


def normalize_clasif(val: str) -> str:
    v = str(val).strip()
    return CAT_MAP.get(v, v)

# ============================================================
# Read one saldo table from a .mdb connection
# Returns DataFrame with columns: [clasif, operacion, tasa, moneda, saldo_usd]
# is_atrasos=True means table has per-payment rows; sum dolares by group
# ============================================================

def read_saldo_table(conn, table_name: str, is_atrasos: bool = False) -> pd.DataFrame | None:
    cursor = conn.cursor()
    cursor.execute(f"SELECT * FROM [{table_name}] WHERE 1=0")
    actual_cols = [d[0] for d in cursor.description]

    # Prefer Clasificacion over Tipo de Deuda; falls back to all CLASIF_ALIASES
    col_clasif = pick_col(actual_cols, CLASIF_ALIASES)
    col_op     = pick_col(actual_cols, OPERACION_ALIASES)
    col_tasa   = pick_col(actual_cols, TASA_ALIASES)
    col_moneda = pick_col(actual_cols, MONEDA_ALIASES)
    col_saldo  = pick_col(actual_cols, SALDO_USD_ALIASES)

    if col_clasif is None or col_saldo is None:
        print(f"    [WARN] Cannot map required cols in '{table_name}'. cols={actual_cols}")
        return None

    sel = [c for c in [col_clasif, col_op, col_tasa, col_moneda, col_saldo] if c]
    safe = ", ".join(f"[{c}]" for c in sel)
    try:
        df = pd.read_sql(f"SELECT {safe} FROM [{table_name}]", conn)
    except Exception as e:
        print(f"    [WARN] Error reading '{table_name}': {e}")
        return None

    df.columns = sel

    # Rename to standard
    rename = {col_clasif: "clasif", col_saldo: "saldo_usd"}
    if col_op:     rename[col_op]     = "operacion"
    if col_tasa:   rename[col_tasa]   = "tasa"
    if col_moneda: rename[col_moneda] = "moneda"
    df = df.rename(columns=rename)

    df["saldo_usd"] = pd.to_numeric(df["saldo_usd"], errors="coerce").fillna(0)
    df["clasif"] = df["clasif"].fillna("").astype(str).str.strip().map(
        lambda v: normalize_clasif(v)
    )
    if "operacion" not in df.columns:
        df["operacion"] = ""
    if "tasa" not in df.columns:
        df["tasa"] = ""
    if "moneda" not in df.columns:
        df["moneda"] = ""

    df["operacion"] = df["operacion"].fillna("").astype(str).str.strip()
    df["tasa"]      = df["tasa"].fillna("").astype(str).str.strip()
    df["moneda"]    = df["moneda"].fillna("").astype(str).str.strip()

    # For atrasos tables: sum per (clasif, operacion, tasa, moneda)
    if is_atrasos:
        df = df.groupby(["clasif", "operacion", "tasa", "moneda"], as_index=False)["saldo_usd"].sum()

    df = df[df["clasif"] != ""].copy()
    return df[["clasif", "operacion", "tasa", "moneda", "saldo_usd"]]

# ============================================================
# Process one .mdb file -> dict of DataFrames per situacion
# ============================================================

def process_mdb(mdb_path: Path) -> dict[str, pd.DataFrame]:
    conn_str = r'DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};' + f'DBQ={mdb_path};'
    try:
        conn = pyodbc.connect(conn_str)
    except Exception as e:
        print(f"  [WARN] Cannot connect: {e}")
        return {}

    cursor = conn.cursor()
    all_tables = [r.table_name for r in cursor.tables(tableType="TABLE") if "~" not in r.table_name]

    result = {}

    # --- Pago Normal ---
    t_normal = (
        find_table(all_tables, ["NORMAL"],       exclude=["VENCIM","REESTRUC"]) or
        find_table(all_tables, ["SALDOS"],        exclude=["VENCIM","ATRASO"]) or
        find_table(all_tables, ["SALDO"],         exclude=["VENCIM","REESTRUC","CANJE","ATRASO"])
    )
    if t_normal:
        df = read_saldo_table(conn, t_normal)
        if df is not None and not df.empty:
            result["normal"] = df
    else:
        print(f"  [WARN] No pago normal table. Available: {all_tables}")

    # --- Reestructurar / Pago Diferido ---
    t_reestr = (
        find_table(all_tables, ["REESTRUCTURAR"],      exclude=["VENCIM","ATRASO"]) or
        find_table(all_tables, ["NO VENCIDA"],          exclude=["VENCIM","ATRASO"])
    )
    if t_reestr:
        # Check if it's an atrasos table (has mora-related columns)
        cursor.execute(f"SELECT * FROM [{t_reestr}] WHERE 1=0")
        acols = [d[0].upper() for d in cursor.description]
        is_atr = any("MORA" in c or "FECHA DE SERVICIO" in c for c in acols) and \
                 not any("SALDO EN" in c for c in acols)
        df = read_saldo_table(conn, t_reestr, is_atrasos=is_atr)
        if df is not None and not df.empty:
            result["reestructurar"] = df
    # If no reestructurar table, just skip (normal for 2016+ after holdout settlement)

    # --- No Presentada al Canje / Elegible Pendiente ---
    # 2007-2015: table named "DEUDA NO PRESENTADA AL CANJE VENCIDA"  -> matched by "CANJE"
    # 2016-2018: table named "Deuda elegible pendiente de reestructuracion" -> matched by "ELEGIBLE"
    t_canje = (
        find_table(all_tables, ["CANJE"],    exclude=["VENCIM", "NO VENCIDA", "ATRASO", "REESTRUC", "NORMAL"]) or
        find_table(all_tables, ["ELEGIBLE"], exclude=["VENCIM", "ATRASO", "NORMAL"])
    )
    if t_canje:
        cursor.execute(f"SELECT * FROM [{t_canje}] WHERE 1=0")
        acols = [d[0].upper() for d in cursor.description]
        is_atr = any("MORA" in c or "TIPO MORA" in c for c in acols) and \
                 not any("SALDO EN" in c for c in acols)
        df = read_saldo_table(conn, t_canje, is_atrasos=is_atr)
        if df is not None and not df.empty:
            result["no_canje"] = df

    conn.close()
    return result

# ============================================================
# Build A.1 structure
# ============================================================

SITUACION_LABELS = {
    "normal":       "I.   DEUDA EN SITUACION DE PAGO NORMAL",
    "reestructurar":"II.  DEUDA EN SITUACION DE PAGO DIFERIDO",
    "no_canje":     "III. DEUDA ELEGIBLE PENDIENTE DE REESTRUCTURACION",
}

# ---- Instrument group membership ----
_TITULOS_CATS = frozenset({"BONOS"})
_LETRAS_CATS  = frozenset({
    "LETRAS DEL TESORO",
    "LETRAS EN GARANTIA", "LETRAS EN GARANTÍA",
})
_ORG_INT_CATS = frozenset({
    "BIRF", "BID", "CAF", "BEI", "FONPLATA", "FIDA", "FMI", "BCIE", "OFID",
})
_ORG_OF_CATS  = frozenset({
    "BILATERALES", "CLUB DE PARIS",
    "ORGANISM. OFICIALES - BILATERAL EXTERNA",
    "ORGANISM. OFICIALES - BILATERAL INTERNA",
})

_CAT_DISPLAY = {
    "BANCA":   "BANCA COMERCIAL",
    "PAGARES": "PAGARES DEL TESORO",
}

_PREST_MAIN_ORDER = [
    "PRESTAMOS GARANTIZADOS",
    "BANCA", "PAGARES", "ANTICIPOS BCRA",
    "DECRETO 977", "DECRETO 1579", "PRESTAMOS", "OTROS",
]


def _cat_group(cl: str) -> str:
    if cl in _TITULOS_CATS:
        return "titulos"
    if cl in _LETRAS_CATS:
        return "letras"
    return "prestamos"


def _prestamos_subgroup(cl: str) -> str | None:
    if cl in _ORG_INT_CATS:
        return "org_int"
    if cl in _ORG_OF_CATS:
        return "org_of"
    return None


def build_a1(all_data: dict[str, dict[str, pd.DataFrame]]) -> pd.DataFrame:
    """
    all_data: { col_label -> { situacion -> DataFrame } }
    Returns DataFrame with hierarchical row labels as index.
    Rows are grouped into TITULOS PUBLICOS / LETRAS / PRESTAMOS per situation.
    """
    from collections import defaultdict

    rows_set = []
    for col_data in all_data.values():
        for sit, df in col_data.items():
            for _, row in df.iterrows():
                cl = row["clasif"]
                op = row["operacion"]
                sc = get_subcategory(cl, op)
                rows_set.append((sit, cl, sc))

    sit_order = ["normal", "reestructurar", "no_canje"]
    sit_clasif: dict[str, set] = defaultdict(set)
    clasif_subcat: dict[tuple, set] = defaultdict(set)
    for sit, cl, sc in rows_set:
        sit_clasif[sit].add(cl)
        if sc:
            clasif_subcat[(sit, cl)].add(sc)

    row_defs = []
    row_defs.append(("DEUDA BRUTA TOTAL (I + II + III)", ("TOTAL", None, None)))

    for sit in sit_order:
        if sit not in sit_clasif:
            continue
        row_defs.append((SITUACION_LABELS[sit], (sit, None, None)))

        all_cl = sit_clasif[sit]
        titulos_cl   = sorted(cl for cl in all_cl if _cat_group(cl) == "titulos")
        letras_cl    = sorted(cl for cl in all_cl if _cat_group(cl) == "letras")
        prestamos_cl = sorted(cl for cl in all_cl if _cat_group(cl) == "prestamos")

        # ---- TITULOS PUBLICOS ----
        if titulos_cl:
            row_defs.append(("  TITULOS PUBLICOS", (sit, "__TITULOS__", None)))
            for cl in titulos_cl:
                disp = _CAT_DISPLAY.get(cl, cl)
                row_defs.append((f"    {disp}", (sit, cl, None)))
                for sc in sorted(clasif_subcat.get((sit, cl), set())):
                    row_defs.append((f"        {sc}", (sit, cl, sc)))

        # ---- LETRAS ----
        if letras_cl:
            row_defs.append(("  LETRAS", (sit, "__LETRAS__", None)))
            for cl in letras_cl:
                disp = _CAT_DISPLAY.get(cl, cl)
                row_defs.append((f"    {disp}", (sit, cl, None)))
                for sc in sorted(clasif_subcat.get((sit, cl), set())):
                    row_defs.append((f"        {sc}", (sit, cl, sc)))

        # ---- PRESTAMOS ----
        if prestamos_cl:
            row_defs.append(("  PRESTAMOS", (sit, "__PRESTAMOS__", None)))

            prest_guar  = [cl for cl in prestamos_cl if cl == "PRESTAMOS GARANTIZADOS"]
            org_int_cl  = sorted(cl for cl in prestamos_cl if _prestamos_subgroup(cl) == "org_int")
            org_of_cl   = sorted(cl for cl in prestamos_cl if _prestamos_subgroup(cl) == "org_of")
            other_prest = [cl for cl in prestamos_cl
                           if cl != "PRESTAMOS GARANTIZADOS"
                           and _prestamos_subgroup(cl) is None]

            def _prest_sort(cl):
                try:
                    return _PREST_MAIN_ORDER.index(cl)
                except ValueError:
                    return len(_PREST_MAIN_ORDER)

            for cl in prest_guar:
                disp = _CAT_DISPLAY.get(cl, cl)
                row_defs.append((f"    {disp}", (sit, cl, None)))
                for sc in sorted(clasif_subcat.get((sit, cl), set())):
                    row_defs.append((f"        {sc}", (sit, cl, sc)))

            if org_int_cl:
                row_defs.append(("    ORGANISMOS INTERNACIONALES", (sit, "__ORG_INT__", None)))
                for cl in org_int_cl:
                    disp = _CAT_DISPLAY.get(cl, cl)
                    row_defs.append((f"      {disp}", (sit, cl, None)))
                    for sc in sorted(clasif_subcat.get((sit, cl), set())):
                        row_defs.append((f"          {sc}", (sit, cl, sc)))

            if org_of_cl:
                row_defs.append(("    ORGANISMOS OFICIALES", (sit, "__ORG_OF__", None)))
                for cl in org_of_cl:
                    disp = _CAT_DISPLAY.get(cl, cl)
                    row_defs.append((f"      {disp}", (sit, cl, None)))
                    for sc in sorted(clasif_subcat.get((sit, cl), set())):
                        row_defs.append((f"          {sc}", (sit, cl, sc)))

            for cl in sorted(other_prest, key=_prest_sort):
                disp = _CAT_DISPLAY.get(cl, cl)
                row_defs.append((f"    {disp}", (sit, cl, None)))
                for sc in sorted(clasif_subcat.get((sit, cl), set())):
                    row_defs.append((f"        {sc}", (sit, cl, sc)))

    def make_lookup(col_data: dict[str, pd.DataFrame]) -> dict:
        lkp = {}
        for sit, df in col_data.items():
            for _, row in df.iterrows():
                cl  = row["clasif"]
                op  = row["operacion"]
                sc  = get_subcategory(cl, op)
                v   = row["saldo_usd"]
                grp = _cat_group(cl)
                psg = _prestamos_subgroup(cl)

                lkp[(sit, cl, None)]       = lkp.get((sit, cl, None), 0) + v
                if sc:
                    lkp[(sit, cl, sc)]     = lkp.get((sit, cl, sc), 0) + v
                lkp[(sit, None, None)]     = lkp.get((sit, None, None), 0) + v
                lkp[("TOTAL", None, None)] = lkp.get(("TOTAL", None, None), 0) + v

                if grp == "titulos":
                    lkp[(sit, "__TITULOS__", None)]   = lkp.get((sit, "__TITULOS__", None), 0) + v
                elif grp == "letras":
                    lkp[(sit, "__LETRAS__", None)]    = lkp.get((sit, "__LETRAS__", None), 0) + v
                else:
                    lkp[(sit, "__PRESTAMOS__", None)] = lkp.get((sit, "__PRESTAMOS__", None), 0) + v
                    if psg == "org_int":
                        lkp[(sit, "__ORG_INT__", None)] = lkp.get((sit, "__ORG_INT__", None), 0) + v
                    elif psg == "org_of":
                        lkp[(sit, "__ORG_OF__", None)]  = lkp.get((sit, "__ORG_OF__", None), 0) + v
        return lkp

    cols = sorted(all_data.keys(), key=col_to_date)
    labels = [r[0] for r in row_defs]
    keys   = [r[1] for r in row_defs]

    data_matrix = {col: [0.0] * len(labels) for col in cols}
    for col in cols:
        col_data = all_data.get(col, {})
        lkp = make_lookup(col_data)
        for i, key in enumerate(keys):
            data_matrix[col][i] = lkp.get(key, 0.0)

    result = pd.DataFrame(data_matrix, index=range(len(labels)), columns=cols, dtype=float)
    result.insert(0, "", labels)
    result = result.set_index("")
    result.index.name = ""
    return result / 1_000_000

# ============================================================
# Build A.3 structure
# ============================================================

def build_a3(all_data: dict[str, dict[str, pd.DataFrame]]) -> pd.DataFrame:
    """
    Uses only 'normal' situation data.
    Rows: (moneda_group1, moneda_group2, tasa_group) hierarchy
    """
    # Collect all (mg1, mg2, tg) combos
    combos = set()
    for col_data in all_data.values():
        df = col_data.get("normal")
        if df is None or df.empty:
            continue
        for _, row in df.iterrows():
            mg1, mg2 = moneda_group(row["moneda"])
            tg = tasa_group(row["tasa"])
            combos.add((mg1, mg2, tg))

    # Fixed ordering
    mg1_order = ["Moneda local", "Moneda extranjera"]
    mg2_order  = {
        "Moneda local":     ["ARP (no ajustable CER)", "UCP (ajustable CER)"],
        "Moneda extranjera":["Dolares", "Euros", "Yenes", "DEG", "Otras monedas"],
    }
    tg_order = ["Tasa Fija", "Tasa Cero", "Tasa Variable"]

    row_defs = []
    row_defs.append(("DEUDA BRUTA EN SITUACION DE PAGO NORMAL", ("TOTAL", None, None)))
    for mg1 in mg1_order:
        row_defs.append((f"  {mg1}", (mg1, None, None)))
        for mg2 in mg2_order[mg1]:
            if not any(c[0] == mg1 and c[1] == mg2 for c in combos):
                continue
            row_defs.append((f"    {mg2}", (mg1, mg2, None)))
            for tg in tg_order:
                if (mg1, mg2, tg) in combos:
                    row_defs.append((f"        {tg}", (mg1, mg2, tg)))

    # Composition by moneda (summary)
    row_defs.append(("", None))
    row_defs.append(("COMPOSICION POR MONEDA", None))
    row_defs.append(("  DEUDA BRUTA EN SITUACION DE PAGO NORMAL", ("TOTAL", None, None)))
    for mg1 in mg1_order:
        for mg2 in mg2_order[mg1]:
            if any(c[0] == mg1 and c[1] == mg2 for c in combos):
                row_defs.append((f"    {mg2}", (mg1, mg2, None)))

    # Composition by tasa (summary)
    row_defs.append(("", None))
    row_defs.append(("COMPOSICION POR TASA", None))
    row_defs.append(("  DEUDA BRUTA EN SITUACION DE PAGO NORMAL", ("TOTAL", None, None)))
    for tg in tg_order:
        row_defs.append((f"    {tg}", ("TASA", None, tg)))

    def make_lookup_a3(df: pd.DataFrame) -> dict:
        lkp = {}
        for _, row in df.iterrows():
            mg1, mg2 = moneda_group(row["moneda"])
            tg = tasa_group(row["tasa"])
            v  = row["saldo_usd"]
            lkp[("TOTAL", None, None)] = lkp.get(("TOTAL", None, None), 0) + v
            lkp[(mg1, None, None)]     = lkp.get((mg1, None, None), 0) + v
            lkp[(mg1, mg2, None)]      = lkp.get((mg1, mg2, None), 0) + v
            lkp[(mg1, mg2, tg)]        = lkp.get((mg1, mg2, tg), 0) + v
            lkp[("TASA", None, tg)]    = lkp.get(("TASA", None, tg), 0) + v
        return lkp

    cols = sorted(all_data.keys(), key=col_to_date)
    labels = [r[0] for r in row_defs]
    keys   = [r[1] for r in row_defs]

    data_matrix = {col: [None] * len(labels) for col in cols}
    for col in cols:
        df = all_data.get(col, {}).get("normal")
        lkp = make_lookup_a3(df) if (df is not None and not df.empty) else {}
        for i, key in enumerate(keys):
            if key is not None:
                data_matrix[col][i] = lkp.get(key, 0.0)

    result = pd.DataFrame(data_matrix, index=range(len(labels)), columns=cols, dtype=float)
    result.insert(0, "", labels)
    result = result.set_index("")
    result.index.name = ""

    return result / 1_000_000

# ============================================================
# Write Excel with boletín-style formatting
# ============================================================

def write_excel(a1: pd.DataFrame, a3: pd.DataFrame, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(str(path), engine="openpyxl") as writer:
        _write_sheet(writer, "A.1", a1,
                     title="DEUDA BRUTA DE LA ADMINISTRACION CENTRAL",
                     subtitle="POR INSTRUMENTO Y SITUACION DE PAGO",
                     note="Datos en millones de U$S")
        _write_sheet(writer, "A.3", a3,
                     title="DEUDA BRUTA DE LA ADMINISTRACION CENTRAL EN SITUACION DE PAGO NORMAL",
                     subtitle="POR MONEDA Y TASA",
                     note="Datos en millones de U$S")


def _write_sheet(writer, sheet_name: str, df: pd.DataFrame,
                 title: str, subtitle: str, note: str) -> None:
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    ws = writer.book.create_sheet(sheet_name)

    # Header block
    ws.cell(1, 2, "MINISTERIO DE ECONOMIA")
    ws.cell(2, 2, "SECRETARIA DE FINANZAS")
    ws.cell(4, 2, title)
    ws.cell(5, 2, subtitle)
    ws.cell(6, 2, note)

    # Column headers (dates) starting row 8, col 2 onwards
    HDR_ROW = 8
    DATA_START_ROW = 9
    LABEL_COL = 1

    cols = list(df.columns)
    for j, col in enumerate(cols):
        cell = ws.cell(HDR_ROW, 2 + j, col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for i, (label, row_data) in enumerate(df.iterrows()):
        r = DATA_START_ROW + i
        ws.cell(r, LABEL_COL, str(label))

        lbl_str      = str(label)
        lbl_stripped = lbl_str.strip()
        leading_sp   = len(lbl_str) - len(lbl_str.lstrip(" "))
        is_bold = (
            leading_sp == 0 or
            leading_sp == 2 or
            "TOTAL" in lbl_stripped or
            "COMPOSICION" in lbl_stripped or
            "DEUDA BRUTA" in lbl_stripped or
            (leading_sp == 4 and lbl_stripped in (
                "ORGANISMOS INTERNACIONALES", "ORGANISMOS OFICIALES",
            ))
        )
        ws.cell(r, LABEL_COL).font = Font(bold=is_bold)

        for j, col in enumerate(cols):
            val = row_data[col]
            if pd.notna(val) and val != 0:
                cell = ws.cell(r, 2 + j, round(float(val), 1))
                cell.number_format = "#,##0.0"

    # Column widths
    ws.column_dimensions[get_column_letter(LABEL_COL)].width = 55
    for j in range(len(cols)):
        ws.column_dimensions[get_column_letter(2 + j)].width = 12

    # Freeze panes
    ws.freeze_panes = ws.cell(DATA_START_ROW, 2)

# ============================================================
# Main
# ============================================================

def main():
    mdb_files = sorted(MDB_DIR.glob("*.mdb"))
    print(f"Found {len(mdb_files)} .mdb files")

    all_data: dict[str, dict[str, pd.DataFrame]] = {}

    for mdb_path in mdb_files:
        file_date = parse_file_date(mdb_path.stem)
        if file_date is None:
            print(f"  [SKIP] Cannot parse date: {mdb_path.name}")
            continue

        col_label = file_date.strftime("%m/%Y")
        print(f"Processing {mdb_path.name}  ->  {col_label}")

        data = process_mdb(mdb_path)
        if data:
            situations = {k: v.shape[0] for k, v in data.items()}
            print(f"    loaded: {situations}")
            all_data[col_label] = data
        else:
            print(f"    [WARN] No data extracted")

    print("\nBuilding A.1...")
    a1 = build_a1(all_data)
    print(f"  A.1 shape: {a1.shape}")

    print("Building A.3...")
    a3 = build_a3(all_data)
    print(f"  A.3 shape: {a3.shape}")

    print(f"Writing to {OUTPUT_PATH}...")
    write_excel(a1, a3, OUTPUT_PATH)
    print("Done.")

    # Quick sanity check: total debt per period
    print("\nSanity check - DEUDA BRUTA TOTAL (millones USD):")
    total_row = a1.loc["DEUDA BRUTA TOTAL (I + II + III)"]
    print(total_row.to_string())


if __name__ == "__main__":
    main()
