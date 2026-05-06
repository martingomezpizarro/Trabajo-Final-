"""
build_graficos_deuda.py

Post-procesa saldo_deuda_sigade.xlsx para producir hojas listas para graficar,
replicando la estructura de los 8 graficos de:
  https://www.argentina.gob.ar/economia/finanzas/graficos-deuda

Hojas de salida en graficos_deuda.xlsx:
  Evolucion       - Deuda total y por situacion de pago (millon USD + %)
  Por_Acreedor    - Sector Publico / Org. Internacionales / Bilateral / Sector Privado
  Por_Moneda      - Detalle por moneda (deuda en pago normal)
  Por_Tasa        - Fija / Variable / Cero (deuda en pago normal)

Nota: LETRAS DEL TESORO se clasifica como Sector Publico Nacional
(en 2007-2018 estaban mayormente en manos de ANSES/FGS/BCRA).
"""

import re
import warnings
from pathlib import Path
from datetime import date

import pandas as pd
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore", category=UserWarning)

INPUT_PATH  = Path(r"c:\Users\Usuario\Desktop\MARTIN\ECONOMICS\TRABAJO FINAL\Trabajo\data\Variables Finales\saldo_deuda.xlsx")
OUTPUT_PATH = Path(r"c:\Users\Usuario\Desktop\MARTIN\ECONOMICS\TRABAJO FINAL\Trabajo\data\Variables Finales\graficos_deuda.xlsx")

# Clasificacion -> tipo de acreedor
ACREEDOR_MAP = {
    # Sector Publico Nacional
    "ANTICIPOS BCRA":                         "Sector Publico Nacional",
    "LETRAS DEL TESORO":                      "Sector Publico Nacional",
    "DECRETO 977":                            "Sector Publico Nacional",
    "DECRETO 1579":                           "Sector Publico Nacional",
    "LETRAS EN GARANT\xeda":                  "Sector Publico Nacional",
    "LETRAS EN GARANTIA":                     "Sector Publico Nacional",
    "LETRAS EN GARANTía":                "Sector Publico Nacional",
    # Organismos Internacionales
    "BID":                                    "Organismos Internacionales",
    "BIRF":                                   "Organismos Internacionales",
    "CAF":                                    "Organismos Internacionales",
    "BEI":                                    "Organismos Internacionales",
    "FONPLATA":                               "Organismos Internacionales",
    "FIDA":                                   "Organismos Internacionales",
    "FMI":                                    "Organismos Internacionales",
    "BCIE":                                   "Organismos Internacionales",
    "OFID":                                   "Organismos Internacionales",
    # Bilateral
    "BILATERALES":                            "Bilateral",
    "CLUB DE PARIS":                          "Bilateral",
    "ORGANISM. OFICIALES - BILATERAL EXTERNA":"Bilateral",
    "ORGANISMOS OFICIALES":                   "Bilateral",    # sub-group header in A.1
    # Sector Privado
    "BONOS":                                  "Sector Privado",
    "BANCA":                                  "Sector Privado",
    "BANCA COMERCIAL":                        "Sector Privado",   # display name in A.1
    "PRESTAMOS GARANTIZADOS":                 "Sector Privado",
    "PRESTAMOS":                              "Sector Privado",
    "PAGARES":                                "Sector Privado",
    "PAGARES DEL TESORO":                     "Sector Privado",   # display name in A.1
    "OTROS":                                  "Sector Privado",
    # Org. Internacionales sub-group header
    "ORGANISMOS INTERNACIONALES":             "Organismos Internacionales",
}

ACREEDOR_ORDER = [
    "Sector Publico Nacional",
    "Organismos Internacionales",
    "Bilateral",
    "Sector Privado",
    "Sin clasificar",
]

# Keys normalized to lowercase + no accents for fuzzy matching
_ACREEDOR_NORM = {
    k.lower().replace("á","a").replace("é","e").replace("í","i")
              .replace("ó","o").replace("ú","u").replace("ñ","n")
              .replace("\xcd","i").replace("\xed","i").replace("\xf3","o")
              .replace("\xfa","u").replace("\xe1","a").replace("\xe9","e"): v
    for k, v in ACREEDOR_MAP.items()
}

def get_acreedor(cat: str) -> str:
    if cat in ACREEDOR_MAP:
        return ACREEDOR_MAP[cat]
    norm = (cat.lower()
              .replace("á","a").replace("é","e").replace("í","i")
              .replace("ó","o").replace("ú","u").replace("ñ","n")
              .replace("\xcd","i").replace("\xed","i").replace("\xf3","o")
              .replace("\xfa","u").replace("\xe1","a").replace("\xe9","e"))
    return _ACREEDOR_NORM.get(norm, "Sin clasificar")

# -----------------------------------------------------------------------

def col_to_date(c) -> date:
    s = str(c).strip()
    mm, yyyy = s.split("/")
    return date(int(yyyy), int(mm), 1)


def normalize_col(c) -> str:
    if isinstance(c, pd.Timestamp):
        return c.strftime("%m/%Y")
    if hasattr(c, "year") and hasattr(c, "month"):
        return f"{c.month:02d}/{c.year}"
    s = str(c).strip()
    m = re.match(r"^(\d{4})-(\d{2})-\d{2}$", s)
    if m:
        return f"{m.group(2)}/{m.group(1)}"
    return s


def read_sheet(path: Path, sheet_name: str) -> pd.DataFrame:
    df = pd.read_excel(path, sheet_name=sheet_name, skiprows=7, index_col=0)
    df.columns = [normalize_col(c) for c in df.columns]
    date_cols = [c for c in df.columns if re.match(r"^\d{2}/\d{4}$", str(c))]
    df = df[date_cols]
    df = df[df.index.notna()]
    df.index = df.index.astype(str)
    df = df[df.index.str.strip() != ""]
    df = df[df.index.str.lower() != "nan"]
    df = df.apply(pd.to_numeric, errors="coerce").fillna(0)
    df = df[sorted(date_cols, key=col_to_date)]
    return df

# -----------------------------------------------------------------------
# Parsing A.1
# -----------------------------------------------------------------------

def parse_a1(a1: pd.DataFrame):
    """
    Iterates A.1 rows in document order to extract:
      grand_total    - Series (date -> millon USD)
      section_totals - {section_key -> Series}
      cat_rows       - {(section_key, clasif_stripped) -> Series}
    """
    grand_total    = None
    section_totals = {}
    cat_rows       = {}
    current_section = None

    for label, row in a1.iterrows():
        s = str(label)

        if "DEUDA BRUTA TOTAL" in s:
            grand_total = row.astype(float)
            continue

        if re.match(r"^I\b", s.strip()) and "NORMAL" in s.upper():
            current_section = "normal"
            section_totals["normal"] = row.astype(float)
            continue
        if re.match(r"^II\b", s.strip()):
            current_section = "reestructurar"
            section_totals["reestructurar"] = row.astype(float)
            continue
        if re.match(r"^III\b", s.strip()):
            current_section = "no_canje"
            section_totals["no_canje"] = row.astype(float)
            continue

        # 4-space rows are individual categories or sub-group headers
        # (ORGANISMOS INTERNACIONALES / ORGANISMOS OFICIALES).
        # 2-space rows are top-level group headers (TITULOS PUBLICOS / LETRAS / PRESTAMOS)
        # — skip those.  6-space and deeper are sub-members and subcats — skip those too.
        if s.startswith("    ") and not s.startswith("      ") and current_section:
            cat = s.strip()
            cat_rows[(current_section, cat)] = row.astype(float)

    return grand_total, section_totals, cat_rows

# -----------------------------------------------------------------------
# Build composition DataFrames
# -----------------------------------------------------------------------

def build_evolucion(grand_total, section_totals) -> pd.DataFrame:
    labels = {
        "normal":        "Pago Normal",
        "reestructurar": "Pago Diferido",
        "no_canje":      "Elegible Pendiente",
    }
    rows = {"Deuda Bruta Total": grand_total}
    for k in ["normal", "reestructurar", "no_canje"]:
        if k in section_totals:
            rows[labels[k]] = section_totals[k]
    return pd.DataFrame(rows).T


def build_por_acreedor(grand_total, cat_rows) -> pd.DataFrame:
    acreedor: dict[str, pd.Series] = {}
    for (section, cat), row in cat_rows.items():
        ac = get_acreedor(cat)
        if ac == "Sin clasificar":
            print(f"  [WARN] Unclassified: '{cat}'")
        acreedor[ac] = acreedor[ac] + row if ac in acreedor else row.copy()

    ordered = {k: acreedor[k] for k in ACREEDOR_ORDER if k in acreedor}
    ordered["Total"] = grand_total

    df = pd.DataFrame(ordered).T
    final_order = ["Total"] + [k for k in ACREEDOR_ORDER if k in df.index]
    return df.reindex([o for o in final_order if o in df.index])


def build_por_moneda_tasa(a3: pd.DataFrame):
    """Returns (df_moneda, df_tasa) from A.3 COMPOSICION sections."""
    moneda_rows: dict[str, pd.Series] = {}
    tasa_rows:   dict[str, pd.Series] = {}
    in_moneda = False
    in_tasa   = False

    for label, row in a3.iterrows():
        s = str(label)
        stripped = s.strip()

        if "COMPOSICION POR MONEDA" in s.upper():
            in_moneda, in_tasa = True, False
            continue
        if "COMPOSICION POR TASA" in s.upper():
            in_tasa, in_moneda = True, False
            continue

        if in_moneda and stripped and stripped.lower() not in ("", "nan"):
            if "DEUDA BRUTA" in stripped.upper():
                moneda_rows["Total Pago Normal"] = row.astype(float)
            else:
                moneda_rows[stripped] = row.astype(float)

        if in_tasa and stripped and stripped.lower() not in ("", "nan"):
            if "DEUDA BRUTA" in stripped.upper():
                tasa_rows["Total Pago Normal"] = row.astype(float)
            else:
                tasa_rows[stripped] = row.astype(float)

    # Fallback: read main section if COMPOSICION sections absent
    if not moneda_rows:
        for label, row in a3.iterrows():
            s = str(label).strip()
            if "DEUDA BRUTA EN SITUACION DE PAGO NORMAL" in s.upper() and not s.startswith(" "):
                moneda_rows["Total Pago Normal"] = row.astype(float)
            elif s in ("Moneda local", "Moneda extranjera"):
                moneda_rows[s] = row.astype(float)

    df_mon = pd.DataFrame(moneda_rows).T
    df_tas = pd.DataFrame(tasa_rows).T
    return df_mon, df_tas


def pct_of_total(df: pd.DataFrame, total_label: str) -> pd.DataFrame:
    if total_label not in df.index:
        return pd.DataFrame(index=df.index, columns=df.columns).fillna(0)
    total = df.loc[total_label].replace(0, float("nan"))
    return (df.div(total) * 100).fillna(0)

# -----------------------------------------------------------------------
# Excel writing helpers
# -----------------------------------------------------------------------

def _write_block(ws, df: pd.DataFrame, start_row: int, fmt: str) -> None:
    cols = list(df.columns)
    # Column header row
    for j, col in enumerate(cols):
        cell = ws.cell(start_row, 2 + j, col)
        cell.font = Font(bold=True, size=8)
        cell.alignment = Alignment(horizontal="center", textRotation=90)
    # Data rows
    for i, (label, row_data) in enumerate(df.iterrows()):
        r = start_row + 1 + i
        s = str(label)
        bold = ("total" in s.lower() or s == s.upper() and len(s) > 2)
        ws.cell(r, 1, s).font = Font(bold=bold)
        for j, col in enumerate(cols):
            val = row_data.get(col, 0) if col in row_data.index else 0
            if pd.notna(val) and val != 0:
                cell = ws.cell(r, 2 + j, round(float(val), 1))
                cell.number_format = fmt


def write_chart_sheet(writer, name: str,
                       df_abs: pd.DataFrame, df_pct: pd.DataFrame,
                       title: str) -> None:
    ws = writer.book.create_sheet(name)
    cols = list(df_abs.columns)

    # Title
    c = ws.cell(1, 1, title)
    c.font = Font(bold=True, size=11)

    # Absolute block
    ws.cell(3, 1, "Millones USD").font = Font(bold=True)
    _write_block(ws, df_abs, start_row=4, fmt="#,##0.0")

    # Percentage block
    pct_start = 4 + len(df_abs) + 2
    ws.cell(pct_start, 1, "% del total").font = Font(bold=True)
    _write_block(ws, df_pct, start_row=pct_start + 1, fmt="#,##0.0")

    # Column widths
    ws.column_dimensions[get_column_letter(1)].width = 34
    for j in range(len(cols)):
        ws.column_dimensions[get_column_letter(2 + j)].width = 9

    ws.freeze_panes = ws.cell(5, 2)

# -----------------------------------------------------------------------
# Main
# -----------------------------------------------------------------------

def main():
    print(f"Reading {INPUT_PATH.name}...")
    a1 = read_sheet(INPUT_PATH, "A.1")
    a3 = read_sheet(INPUT_PATH, "A.3")
    print(f"  A.1: {a1.shape}   A.3: {a3.shape}")

    grand_total, section_totals, cat_rows = parse_a1(a1)
    if grand_total is None:
        print("[ERROR] 'DEUDA BRUTA TOTAL' row not found in A.1")
        return
    print(f"  Sections parsed: {list(section_totals.keys())}")
    unique_cats = set(c for (_, c) in cat_rows)
    print(f"  Unique categories: {sorted(unique_cats)}")

    # Build tables
    evolucion    = build_evolucion(grand_total, section_totals)
    por_acreedor = build_por_acreedor(grand_total, cat_rows)
    df_moneda, df_tasa = build_por_moneda_tasa(a3)

    pct_evolucion = pct_of_total(evolucion,    "Deuda Bruta Total")
    pct_acreedor  = pct_of_total(por_acreedor,  "Total")
    pct_moneda    = pct_of_total(df_moneda,     "Total Pago Normal")
    pct_tasa      = pct_of_total(df_tasa,       "Total Pago Normal")

    print(f"\nEvolucion rows:    {list(evolucion.index)}")
    print(f"Por_Acreedor rows: {list(por_acreedor.index)}")
    print(f"Por_Moneda rows:   {list(df_moneda.index)}")
    print(f"Por_Tasa rows:     {list(df_tasa.index)}")

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(str(OUTPUT_PATH), engine="openpyxl") as writer:
        writer.book.create_sheet("_tmp")
        write_chart_sheet(writer, "Evolucion",
                          evolucion, pct_evolucion,
                          "Evolucion Deuda Bruta - Por Situacion de Pago (millones USD)")
        write_chart_sheet(writer, "Por_Acreedor",
                          por_acreedor, pct_acreedor,
                          "Composicion por Tipo de Acreedor (millones USD)")
        write_chart_sheet(writer, "Por_Moneda",
                          df_moneda, pct_moneda,
                          "Composicion por Moneda - Deuda Pago Normal (millones USD)")
        write_chart_sheet(writer, "Por_Tasa",
                          df_tasa, pct_tasa,
                          "Composicion por Tasa - Deuda Pago Normal (millones USD)")
        del writer.book["_tmp"]

    print(f"\nSaved -> {OUTPUT_PATH}")

    # Sanity prints
    last_col = sorted(grand_total.index, key=col_to_date)[-1]
    first_col = sorted(grand_total.index, key=col_to_date)[0]
    print(f"\nDeuda Bruta Total: {first_col} = {grand_total[first_col]:,.0f}  |  {last_col} = {grand_total[last_col]:,.0f} mill.USD")

    print(f"\nComposicion por acreedor ({last_col}):")
    for label in por_acreedor.index:
        v   = por_acreedor.loc[label, last_col]
        pct = pct_acreedor.loc[label, last_col]
        print(f"  {label:<36} {v:>10,.0f}  {pct:>6.1f}%")

    print(f"\nComposicion por moneda ({last_col}):")
    for label in df_moneda.index:
        v   = df_moneda.loc[label, last_col]
        pct = pct_moneda.loc[label, last_col]
        print(f"  {label:<36} {v:>10,.0f}  {pct:>6.1f}%")

    print(f"\nComposicion por tasa ({last_col}):")
    for label in df_tasa.index:
        v   = df_tasa.loc[label, last_col]
        pct = pct_tasa.loc[label, last_col]
        print(f"  {label:<36} {v:>10,.0f}  {pct:>6.1f}%")


if __name__ == "__main__":
    main()
