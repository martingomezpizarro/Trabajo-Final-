"""
build_vencimientos_2y.py

Serie histórica de vencimientos a 2 años por Tipo de Deuda y Moneda (SIGADE .mdb).
Cobertura: 2007-2025 (todos los archivos .mdb disponibles).

Clasificación de moneda:
  Moneda Local    → código ARP o descripción contiene "PESO"
  Moneda Extranjera → todo lo demás (USD, EUR, JPY, SDR, UCP, etc.)

Todos los valores en millones de USD (columnas Principal en Dólares + Interés en Dólares).

Output: data/Variables Finales/vencimientos_2y.xlsx
  Hoja "Por Tipo y Moneda" — estructura jerárquica:
      TOTAL
        Moneda Local
        Moneda Extranjera
      <Tipo de Deuda 1>
        Moneda Local
        Moneda Extranjera
      <Tipo de Deuda 2>
        ...
  Hoja "Resumen" — totales por tipo (sin desglose moneda), idéntico al output anterior
"""

import re
import warnings
from pathlib import Path
from datetime import date

import pandas as pd
import numpy as np
import pyodbc
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore", category=UserWarning)

MDB_DIR     = Path(r"c:\Users\Usuario\Desktop\MARTIN\ECONOMICS\TRABAJO FINAL\Trabajo\data\basesingade deuda")
OUTPUT_PATH = Path(r"c:\Users\Usuario\Desktop\MARTIN\ECONOMICS\TRABAJO FINAL\Trabajo\data\Variables Finales\vencimientos_2y_nuevo.xlsx")

# ---------------------------------------------------------------------------
# Column alias lists — handles accent/encoding variations across file vintages
# ---------------------------------------------------------------------------

TIPO_ALIASES      = ["Tipo de Deuda", "BOLETIN"]
FECHA_ALIASES     = ["Fecha de Servicio"]
PRINCIPAL_ALIASES = [
    "Principal en Dolares", "Principal en dolares",
    "Principal en D\xf3lares", "Principal en D\xf2lares",
    "Principal en dólares", "Principal en Dólares",
]
INTERES_ALIASES   = [
    "Interes en Dolares", "Interes en dolares",
    "Interes en D\xf3lares", "Interes en D\xf2lares",
    "Interes en dólares", "Interes en Dólares",
]
# 2007-2018: column "Moneda" (ISO codes: USD, ARP, EUR ...)
# 2019-2025: columns "Moneda de Origen" (code) + "Descripción" (name)
MONEDA_ORIG_ALIASES = ["Moneda de Origen", "Moneda"]
MONEDA_DESC_ALIASES = [
    "Descripci\xf3n", "Descripcion", "Descripción",
]

# Codes / substrings that mean Argentine Peso (local currency)
LOCAL_CODES = {"ARP"}
LOCAL_DESC_SUBSTR = "PESO"

# ---------------------------------------------------------------------------
# Canonical name mapping — unifies historical variants into a single label
# ---------------------------------------------------------------------------
TIPO_MAP = {
    # 1. Anticipos BCRA
    "ANTICIPOS DEL BANCO CENTRAL - Cto.Plazo":   "ANTICIPOS BCRA",
    # 2. Banca comercial (externa + interna → una sola fila)
    "BANCA":                                      "BANCA COMERCIAL",
    "BANCA COMERCIAL -BANCA PRIVADA EXTERNA":     "BANCA COMERCIAL",
    "BANCA COMERCIAL-BANCA PRIVADA INTERNA":      "BANCA COMERCIAL",
    # 3. Organismos internacionales — variantes de archivos 2007-2008
    "ORGANISMOS INTERNACIONALES - BEI":           "BEI",
    "ORGANISMOS INTERNACIONALES - BID/FIDA":      "BID",
    "ORGANISMOS INTERNACIONALES - CAF":           "CAF",
    "ORGANISMOS INTERNACIONALES - FIDA":          "FIDA",
    "ORGANISMOS INTERNACIONALES -BIRF":           "BIRF",
    "ORGANISMOS INTERNACIONALES -FONPLATA":       "FONPLATA",
    "BID/FIDA":                                   "BID",
    # Bilaterales (organismos oficiales — distintas grafías históricas)
    "ORGANISM. OFICIALES - BILATERAL EXTERNA":    "BILATERALES",
    "ORGANISM. OFICIALES -BILATERAL EXTERNA":     "BILATERALES",
    # 4. Decreto 977
    "DECRETO 977 - BOGAR 2020":                   "DECRETO 977",
    # 5. Préstamos garantizados (singular y variantes de tasa)
    "PRESTAMO GARANTIZADO TASA VARIABLE":         "PRESTAMOS GARANTIZADOS",
    "PRESTAMO GARANTIZADO TASA FIJA":             "PRESTAMOS GARANTIZADOS",
    "PRESTAMO GARANTIZADO TASA FIJA VTO 2011":    "PRESTAMOS GARANTIZADOS",
    "PRESTAMOS GARANTIZADOS TASA VARIABLE":       "PRESTAMOS GARANTIZADOS",
    "PRESTAMOS GARANTIZADOS TASA FIJA":           "PRESTAMOS GARANTIZADOS",
    "PRESTAMOS GARANTIZADOS TASA FIJA VTO 2011":  "PRESTAMOS GARANTIZADOS",
    # 6. Bonos (títulos públicos con distintos nombres históricos)
    "TITULOS PUBLICOS -Bonos LP":                 "BONOS",
    "TITULOS PUBLICOS -Bonos Pesificados":        "BONOS",
    "TITULOS PUBLICOS - Bonos LP":                "BONOS",
    "TITULOS PUBLICOS - Bonos Pesificados":       "BONOS",
    "TITULOS PUBLICOS - BOCON 9\xa2 SERIE/$/2010/PR16": "BONOS",
    "BOCON 9\xa2 SERIE/$/2010/PR16":             "BONOS",
}

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def parse_file_date(stem: str) -> date | None:
    m = re.search(r"(\d{4})-(\d{2})-(\d{2})$", stem)
    if not m:
        return None
    y, a, b = int(m.group(1)), int(m.group(2)), int(m.group(3))
    if a > 12:
        day, month = a, b
    else:
        month, day = a, b
    try:
        return date(y, month, day)
    except ValueError:
        return None


def find_vencim_table(tables: list[str]) -> str | None:
    for t in tables:
        tu = t.upper()
        if "VENCIM" in tu and "NORMAL" in tu:
            return t
    return None


def pick_col(col_list: list[str], aliases: list[str]) -> str | None:
    col_lower = {c.lower(): c for c in col_list}
    for a in aliases:
        if a in col_list:
            return a
        if a.lower() in col_lower:
            return col_lower[a.lower()]
    return None


def is_local_currency(moneda_code: str, moneda_desc: str) -> bool:
    """Return True if the currency is Argentine Peso (local)."""
    code = str(moneda_code).strip().upper() if moneda_code else ""
    desc = str(moneda_desc).strip().upper() if moneda_desc else ""
    if code in LOCAL_CODES:
        return True
    if LOCAL_DESC_SUBSTR in code or LOCAL_DESC_SUBSTR in desc:
        return True
    return False


def read_vencimientos(mdb_path: Path, file_date: date) -> pd.DataFrame | None:
    """
    Read the vencimientos-normal table from a .mdb file.
    Returns a DataFrame with columns:
        tipo_deuda, moneda_local (bool), principal_usd, interes_usd
    filtered to the 2-year window [file_date, file_date + 2 years].
    """
    conn_str = r'DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};' + f'DBQ={mdb_path};'
    try:
        conn = pyodbc.connect(conn_str)
    except Exception as e:
        print(f"  [WARN] Cannot connect to {mdb_path.name}: {e}")
        return None

    cursor = conn.cursor()
    tables = [r.table_name for r in cursor.tables(tableType="TABLE")]
    vtable = find_vencim_table(tables)

    if vtable is None:
        print(f"  [WARN] No vencimientos table in {mdb_path.name}. Tables: {tables}")
        conn.close()
        return None

    cursor.execute(f"SELECT * FROM [{vtable}] WHERE 1=0")
    actual_cols = [d[0] for d in cursor.description]

    col_tipo  = pick_col(actual_cols, TIPO_ALIASES)
    col_fecha = pick_col(actual_cols, FECHA_ALIASES)
    col_princ = pick_col(actual_cols, PRINCIPAL_ALIASES)
    col_int   = pick_col(actual_cols, INTERES_ALIASES)
    col_mon   = pick_col(actual_cols, MONEDA_ORIG_ALIASES)
    col_desc  = pick_col(actual_cols, MONEDA_DESC_ALIASES)

    missing = [n for n, c in [("TipoDeuda", col_tipo), ("FechaServicio", col_fecha),
                               ("Principal", col_princ), ("Interes", col_int)] if c is None]
    if missing:
        print(f"  [WARN] Missing columns {missing} in {mdb_path.name}. Available: {actual_cols}")
        conn.close()
        return None

    select_cols = [col_tipo, col_fecha, col_princ, col_int]
    if col_mon:
        select_cols.append(col_mon)
    if col_desc:
        select_cols.append(col_desc)

    safe_cols = ", ".join(f"[{c}]" for c in select_cols)
    try:
        df = pd.read_sql(f"SELECT {safe_cols} FROM [{vtable}]", conn)
    except Exception as e:
        print(f"  [WARN] Error reading {vtable} in {mdb_path.name}: {e}")
        conn.close()
        return None
    conn.close()

    # Standardise column names
    df.columns = (
        ["tipo_deuda", "fecha_serv", "principal_usd", "interes_usd"]
        + (["moneda_code"] if col_mon else [])
        + (["moneda_desc"] if col_desc else [])
    )

    df["fecha_serv"]    = pd.to_datetime(df["fecha_serv"], errors="coerce")
    df["principal_usd"] = pd.to_numeric(df["principal_usd"], errors="coerce").fillna(0)
    df["interes_usd"]   = pd.to_numeric(df["interes_usd"],   errors="coerce").fillna(0)

    # Filter: 2-year window
    cutoff = pd.Timestamp(file_date) + pd.DateOffset(years=2)
    mask   = (df["fecha_serv"] >= pd.Timestamp(file_date)) & (df["fecha_serv"] <= cutoff)
    df = df[mask].copy()

    # Classify currency
    moneda_code = df["moneda_code"] if "moneda_code" in df.columns else pd.Series([""] * len(df))
    moneda_desc = df["moneda_desc"] if "moneda_desc" in df.columns else pd.Series([""] * len(df))
    df["moneda_local"] = [
        is_local_currency(c, d)
        for c, d in zip(moneda_code, moneda_desc)
    ]

    return df[["tipo_deuda", "moneda_local", "principal_usd", "interes_usd"]]


# ---------------------------------------------------------------------------
# Aggregation
# ---------------------------------------------------------------------------

def aggregate_records(df: pd.DataFrame) -> dict[tuple, float]:
    """Returns dict: (tipo_deuda, moneda_local) -> total_usd."""
    df = df.copy()
    # Apply canonical name mapping before grouping
    df["tipo_deuda"] = df["tipo_deuda"].map(
        lambda t: TIPO_MAP.get(str(t).strip(), str(t).strip())
    )
    df["total_usd"] = df["principal_usd"] + df["interes_usd"]
    agg = df.groupby(["tipo_deuda", "moneda_local"])["total_usd"].sum()
    return agg.to_dict()


# ---------------------------------------------------------------------------
# Excel output
# ---------------------------------------------------------------------------

FILL_TOTAL     = PatternFill("solid", fgColor="2F5496")   # dark blue — TOTAL rows
FILL_TIPO      = PatternFill("solid", fgColor="D9E1F2")   # light blue — tipo de deuda
FILL_LOCAL     = PatternFill("solid", fgColor="E2EFDA")   # green — Moneda Local
FILL_EXT       = PatternFill("solid", fgColor="FFF2CC")   # yellow — Moneda Extranjera
FILL_HDR       = PatternFill("solid", fgColor="1F3864")   # header

FONT_WHITE_B   = Font(bold=True, color="FFFFFF", size=9)
FONT_BOLD_DARK = Font(bold=True, color="1F3864", size=9)
FONT_NORMAL    = Font(size=9)


def build_row_defs(all_tipos: list[str]) -> list[tuple]:
    """
    Returns list of (label, indent, fill, font, tipo_key, is_local_flag).
    tipo_key=None for TOTAL rows.
    is_local_flag: True=local, False=ext, None=total-of-tipo
    """
    rows = []
    # Grand total block
    rows.append(("TOTAL",              0, FILL_TOTAL, FONT_WHITE_B,   None, None))
    rows.append(("  Moneda Local",     2, FILL_LOCAL, FONT_BOLD_DARK, None, True))
    rows.append(("  Moneda Extranjera",2, FILL_EXT,   FONT_BOLD_DARK, None, False))
    # Per tipo
    for tipo in all_tipos:
        rows.append((tipo,              0, FILL_TIPO,  FONT_BOLD_DARK, tipo, None))
        rows.append(("  Moneda Local",  2, FILL_LOCAL, FONT_NORMAL,    tipo, True))
        rows.append(("  Moneda Extranjera", 2, FILL_EXT, FONT_NORMAL,  tipo, False))
    return rows


def write_detail_sheet(ws, all_tipos: list[str], all_cols: list[str],
                        col_records: dict[str, dict]) -> None:
    """Write the 'Por Tipo y Moneda' sheet."""
    # ---- Header rows ----
    ws.cell(1, 1, "MINISTERIO DE ECONOMIA — SECRETARIA DE FINANZAS")
    ws.cell(2, 1, "VENCIMIENTOS A 2 AÑOS POR TIPO DE DEUDA Y MONEDA")
    ws.cell(3, 1, "SERIE HISTORICA SIGADE 2007-2025 | Millones de USD")
    ws.cell(4, 1, "Moneda Local = Peso Argentino (ARP) | Moneda Extranjera = USD, EUR, JPY, SDR, etc.")

    HDR_ROW    = 6
    DATA_START = 7

    ws.cell(HDR_ROW, 1, "Tipo de Deuda / Moneda").font = Font(bold=True, size=9)
    ws.cell(HDR_ROW, 1).fill = PatternFill("solid", fgColor="1F3864")
    ws.cell(HDR_ROW, 1).font = FONT_WHITE_B

    for j, col in enumerate(all_cols):
        c = ws.cell(HDR_ROW, 2 + j, col)
        c.font      = FONT_WHITE_B
        c.fill      = FILL_HDR
        c.alignment = Alignment(horizontal="center")

    # ---- Row definitions ----
    row_defs = build_row_defs(all_tipos)

    for i, (label, indent, fill, font, tipo_key, is_local) in enumerate(row_defs):
        r = DATA_START + i
        lbl_cell = ws.cell(r, 1, label)
        lbl_cell.fill = fill
        lbl_cell.font = font

        for j, col in enumerate(all_cols):
            records = col_records.get(col, {})
            if tipo_key is None:
                # TOTAL block
                if is_local is None:
                    val = sum(v for (t, loc), v in records.items())
                else:
                    val = sum(v for (t, loc), v in records.items() if loc == is_local)
            else:
                # Specific tipo block
                if is_local is None:
                    val = sum(v for (t, loc), v in records.items() if t == tipo_key)
                else:
                    val = records.get((tipo_key, is_local), 0.0)

            # Convert to millions
            val_m = val / 1e6
            if val_m != 0:
                cell = ws.cell(r, 2 + j, round(val_m, 3))
                cell.number_format = "#,##0.000"
                cell.fill = fill

    # ---- Column widths ----
    ws.column_dimensions["A"].width = 38
    for j in range(len(all_cols)):
        ws.column_dimensions[get_column_letter(2 + j)].width = 11
    ws.freeze_panes = ws.cell(DATA_START, 2)


def write_resumen_sheet(ws, all_tipos: list[str], all_cols: list[str],
                         col_records: dict[str, dict]) -> None:
    """Write the 'Resumen' sheet — totals by tipo only (no moneda breakdown)."""
    ws.cell(1, 1, "MINISTERIO DE ECONOMIA — SECRETARIA DE FINANZAS")
    ws.cell(2, 1, "VENCIMIENTOS A 2 AÑOS POR TIPO DE DEUDA")
    ws.cell(3, 1, "SERIE HISTORICA SIGADE 2007-2025 | Millones de USD")

    HDR_ROW    = 5
    DATA_START = 6

    ws.cell(HDR_ROW, 1, "Tipo de Deuda").font = FONT_WHITE_B
    ws.cell(HDR_ROW, 1).fill = FILL_HDR

    for j, col in enumerate(all_cols):
        c = ws.cell(HDR_ROW, 2 + j, col)
        c.font      = FONT_WHITE_B
        c.fill      = FILL_HDR
        c.alignment = Alignment(horizontal="center")

    # TOTAL row
    ws.cell(DATA_START, 1, "TOTAL").font = FONT_WHITE_B
    ws.cell(DATA_START, 1).fill = FILL_TOTAL
    for j, col in enumerate(all_cols):
        records = col_records.get(col, {})
        val_m   = sum(records.values()) / 1e6
        if val_m:
            cell = ws.cell(DATA_START, 2 + j, round(val_m, 3))
            cell.number_format = "#,##0.000"
            cell.fill = FILL_TOTAL
            cell.font = FONT_WHITE_B

    for i, tipo in enumerate(all_tipos):
        r = DATA_START + 1 + i
        ws.cell(r, 1, tipo).fill = FILL_TIPO
        ws.cell(r, 1).font = FONT_BOLD_DARK
        for j, col in enumerate(all_cols):
            records = col_records.get(col, {})
            val_m   = sum(v for (t, loc), v in records.items() if t == tipo) / 1e6
            if val_m:
                cell = ws.cell(r, 2 + j, round(val_m, 3))
                cell.number_format = "#,##0.000"
                cell.fill = FILL_TIPO

    ws.column_dimensions["A"].width = 38
    for j in range(len(all_cols)):
        ws.column_dimensions[get_column_letter(2 + j)].width = 11
    ws.freeze_panes = ws.cell(DATA_START, 2)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    mdb_files = sorted(MDB_DIR.glob("*.mdb"))
    print(f"Found {len(mdb_files)} .mdb files")

    # col_records: col_label -> {(tipo_deuda, moneda_local): total_usd}
    col_records: dict[str, dict] = {}
    ordered_cols: list[str] = []

    for mdb_path in mdb_files:
        file_date = parse_file_date(mdb_path.stem)
        if file_date is None:
            print(f"  [SKIP] Cannot parse date from: {mdb_path.name}")
            continue

        col_label = file_date.strftime("%m/%Y")
        print(f"Processing {mdb_path.name}  ->  {col_label}", end="  ")

        df = read_vencimientos(mdb_path, file_date)
        if df is None or df.empty:
            col_records[col_label] = {}
            ordered_cols.append(col_label)
            print("(empty)")
            continue

        agg = aggregate_records(df)
        col_records[col_label] = agg
        ordered_cols.append(col_label)

        total = sum(agg.values()) / 1e6
        local = sum(v for (t, loc), v in agg.items() if loc) / 1e6
        ext   = total - local
        print(f"total={total:,.1f} M  local={local:,.1f} M  ext={ext:,.1f} M")

    # Deduplicate columns (keep first occurrence in order)
    seen = set()
    all_cols = []
    for c in ordered_cols:
        if c not in seen:
            all_cols.append(c)
            seen.add(c)

    # All tipo_deuda labels (sorted)
    all_tipos = sorted({
        t for rec in col_records.values() for (t, loc) in rec
    })

    print(f"\nColumns (time periods): {len(all_cols)}  ({all_cols[0]} - {all_cols[-1]})")
    print(f"Tipos de deuda:         {len(all_tipos)}")

    # Write Excel
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    from openpyxl import Workbook
    wb = Workbook()
    wb.remove(wb.active)

    ws1 = wb.create_sheet("Por Tipo y Moneda")
    write_detail_sheet(ws1, all_tipos, all_cols, col_records)

    ws2 = wb.create_sheet("Resumen")
    write_resumen_sheet(ws2, all_tipos, all_cols, col_records)

    wb.save(OUTPUT_PATH)
    print(f"\nSaved -> {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
