"""
build_resultado_fiscal.py

Construye una serie historica mensual y trimestral del resultado fiscal
del Sector Publico Nacional (SPN) / Administracion Publica Nacional (APN)
a partir de los archivos AIF de MECON.

Periodos y fuentes:
  1993-01 a 2006-12: aif_spn_..._mensuales_1993_2006.csv  (SPN)
  2007-01 a 2014-12: aif_spn_..._mensuales_2007_2014.csv  (SPN)
  2015-01 a 2026+:   aif_apn_..._mensuales_2014_2026.csv  (APN, nota: cobertura distinta)

Variables de salida (millones de ARS nominales):
  ingresos_totales       - ingresos corrientes + capital (antes figurativos)
  gtos_primarios         - gastos totales menos intereses (antes figurativos)
  intereses_loc          - intereses en moneda local
  intereses_ext          - intereses en moneda extranjera
  intereses_total        - intereses_loc + intereses_ext
  superavit_primario     - resultado primario
  resultado_financiero   - resultado financiero (incluye intereses)
  emision_loc            - emision de deuda en moneda local
  emision_ext            - emision de deuda en moneda extranjera
  amort_loc              - amortizacion de deuda en moneda local
  amort_ext              - amortizacion de deuda en moneda extranjera

Output: data/processed/resultado_fiscal.xlsx
  Hoja "Mensual"      - valores mensuales
  Hoja "Trimestral"   - acumulado trimestral (suma de los 3 meses)
  Hoja "Anual"        - acumulado anual
"""

import warnings
from pathlib import Path

import pandas as pd
import numpy as np

warnings.filterwarnings("ignore", category=UserWarning)

MECON_DIR   = Path(r"c:\Users\Usuario\Desktop\MARTIN\ECONOMICS\TRABAJO FINAL\Trabajo\data\raw\mecon")
OUTPUT_PATH = Path(r"c:\Users\Usuario\Desktop\MARTIN\ECONOMICS\TRABAJO FINAL\Trabajo\data\processed\resultado_fiscal.xlsx")

# ============================================================
# Column mappings per period
# Each entry: standardized_name -> raw_column_name
# ============================================================

COLS_1993 = {
    "ingresos_totales":   "ing_antes_figurativos_1993_2006",
    "gtos_primarios":     "gtos_primarios_antes_figurativos_1993_2006",
    "intereses_loc":      "gtos_corr_int_ot_ren_prop_int_mon_loc_1993_2006",
    "intereses_ext":      "gtos_corr_int_ot_ren_prop_int_mon_extra_1993_2006",
    "superavit_primario": "superavit_primario_1993_2006",
    "resultado_financiero":"resultado_fin_1993_2006",
    "emision_loc":        "ftes_fin_end_pub_incr_ot_pasiv_end_mon_loc_1993_2006",
    "emision_ext":        "ftes_fin_end_pub_incr_ot_pasiv_end_mon_extra_1993_2006",
    "amort_loc":          "ap_fin_amort_deu_dism_ots_pasiv_amort_mon_loc_1993_2006",
    "amort_ext":          "ap_fin_amort_deu_dism_ots_pasiv_amort_mon_extra_1993_2006",
}

COLS_2007 = {
    "ingresos_totales":   "ing_antes_figurativos_2007_2014",
    "gtos_primarios":     "gtos_primarios_antes_figurativos_2007_2014",
    "intereses_loc":      "gtos_corr_int_ot_ren_prop_int_mon_loc_2007_2014",
    "intereses_ext":      "gtos_corr_int_ot_ren_prop_int_mon_extra_2007_2014",
    "superavit_primario": "superavit_primario_2007_2014",
    "resultado_financiero":"resultado_fin_2007_2014",
    "emision_loc":        "ftes_fin_end_pub_incr_ot_pasiv_end_mon_loc_2007_2014",
    "emision_ext":        "ftes_fin_end_pub_incr_ot_pasiv_end_mon_extra_2007_2014",
    "amort_loc":          "ap_fin_amort_deu_dism_ots_pasiv_amort_mon_loc_2007_2014",
    "amort_ext":          "ap_fin_amort_deu_dism_ots_pasiv_amort_mon_extra_2007_2014",
}

# APN 2015-2026 — cobertura distinta a SPN
COLS_APN = {
    "ingresos_totales":   "ing_antes_figurativos_2017",
    "gtos_primarios":     "gtos_primarios_despues_figurativos_2017",
    "intereses_loc":      "gtos_corr_int_ot_ren_prop_int_mon_loc_2017",
    "intereses_ext":      "gtos_corr_int_ot_ren_prop_int_mon_extra_2017",
    "superavit_primario": "superavit_primario_2017",
    "resultado_financiero":"resultado_fin_2017",
    "emision_loc":        "ftes_fin_end_pub_incr_ot_pasiv_end_mon_loc_2017",
    "emision_ext":        "ftes_fin_end_pub_incr_ot_pasiv_end_mon_extra_2017",
    "amort_loc":          "ap_fin_amort_deu_dism_ots_pasiv_amort_mon_loc_2017",
    "amort_ext":          "ap_fin_amort_deu_dism_ots_pasiv_amort_mon_extra_2017",
}

OUTPUT_COLS = list(COLS_1993.keys())


def read_aif(path: Path, col_map: dict) -> pd.DataFrame:
    """Read one AIF CSV and return standardized DataFrame."""
    raw = pd.read_csv(path, parse_dates=["indice_tiempo"])
    out = pd.DataFrame({"fecha": raw["indice_tiempo"]})
    for std_col, raw_col in col_map.items():
        if raw_col in raw.columns:
            out[std_col] = pd.to_numeric(raw[raw_col], errors="coerce")
        else:
            print(f"  [WARN] Column not found: '{raw_col}' in {path.name}")
            out[std_col] = np.nan
    return out


def aggregate_period(df: pd.DataFrame, freq: str) -> pd.DataFrame:
    """Aggregate to quarterly ('Q') or annual ('Y') by summing."""
    df2 = df.set_index("fecha").copy()
    # Flow variables: sum; only NaN if ALL months in period are NaN
    agg = df2.resample(freq).sum(min_count=1)
    agg.index = agg.index.to_period(freq)
    return agg.reset_index().rename(columns={"fecha": "periodo"})


def format_period_label(p) -> str:
    """Convert Period to display string."""
    try:
        if hasattr(p, "quarter"):
            return f"Q{p.quarter} {p.year}"
        return str(p.year)
    except Exception:
        return str(p)


def write_excel(dfs: dict[str, pd.DataFrame], path: Path) -> None:
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    path.parent.mkdir(parents=True, exist_ok=True)

    FILL_HEADER = PatternFill("solid", fgColor="D9E1F2")

    VALUE_COLS = OUTPUT_COLS
    LABEL_HEADERS = {
        "ingresos_totales":    "Ingresos Totales",
        "gtos_primarios":      "Gastos Primarios",
        "intereses_loc":       "Intereses Mon. Local",
        "intereses_ext":       "Intereses Mon. Extranjera",
        "intereses_total":     "Intereses Total",
        "superavit_primario":  "Resultado Primario",
        "resultado_financiero":"Resultado Financiero",
        "emision_loc":         "Emision Deuda Local",
        "emision_ext":         "Emision Deuda Externa",
        "amort_loc":           "Amortizacion Deuda Local",
        "amort_ext":           "Amortizacion Deuda Externa",
    }

    with pd.ExcelWriter(str(path), engine="openpyxl") as writer:
        for sheet_name, df in dfs.items():
            ws = writer.book.create_sheet(sheet_name)

            ws.cell(1, 1, "MINISTERIO DE ECONOMIA")
            ws.cell(2, 1, "SECRETARIA DE HACIENDA")
            ws.cell(4, 1, "RESULTADO FISCAL DEL SECTOR PUBLICO NACIONAL (SPN/APN)")
            ws.cell(5, 1, "AHORRO, INVERSION Y FINANCIAMIENTO")
            ws.cell(6, 1, "Unidades: millones de ARS nominales")
            ws.cell(7, 1, "Nota: 1993-2014 = SPN; 2015+ = APN (cobertura distinta)")

            HDR_ROW = 9
            DATA_ROW = 10

            # Row headers in col A
            ws.cell(HDR_ROW, 1, "Concepto")
            ws.cell(HDR_ROW, 1).font = Font(bold=True)
            ws.cell(HDR_ROW, 1).fill = FILL_HEADER

            # Time axis: detect period column name
            period_col = df.columns[0]  # "fecha" or "periodo"
            periods = df[period_col].tolist()

            # Column headers (dates/periods)
            for j, p in enumerate(periods):
                if hasattr(p, "strftime"):
                    lbl = p.strftime("%m/%Y")
                else:
                    lbl = format_period_label(p)
                c = ws.cell(HDR_ROW, 2 + j, lbl)
                c.font = Font(bold=True)
                c.alignment = Alignment(horizontal="center")
                c.fill = FILL_HEADER

            # Data rows
            all_vcols = VALUE_COLS + ["intereses_total"]
            for i, vcol in enumerate(all_vcols):
                r = DATA_ROW + i
                lbl = LABEL_HEADERS.get(vcol, vcol)
                is_result = vcol in ("superavit_primario", "resultado_financiero")
                is_int    = vcol == "intereses_total"
                ws.cell(r, 1, lbl).font = Font(bold=(is_result or is_int))

                for j, p in enumerate(periods):
                    if vcol == "intereses_total":
                        val_loc = df.at[j, "intereses_loc"] if "intereses_loc" in df.columns else np.nan
                        val_ext = df.at[j, "intereses_ext"] if "intereses_ext" in df.columns else np.nan
                        val = (val_loc or 0) + (val_ext or 0)
                        if np.isnan(val_loc) and np.isnan(val_ext):
                            val = np.nan
                    else:
                        val = df.at[j, vcol] if vcol in df.columns else np.nan

                    if pd.notna(val):
                        cell = ws.cell(r, 2 + j, round(float(val), 1))
                        cell.number_format = "#,##0.0"
                        if is_result:
                            cell.font = Font(bold=True,
                                            color="375623" if val >= 0 else "C00000")

            # Column widths
            ws.column_dimensions["A"].width = 30
            for j in range(len(periods)):
                ws.column_dimensions[get_column_letter(2 + j)].width = 13

            ws.freeze_panes = ws.cell(DATA_ROW, 2)

    print(f"Saved -> {path}")


def main():
    # --- Load periods ---
    path_1993 = MECON_DIR / "aif_spn_sector_p_blico_nacional_valores_mensuales_1993_2006.csv"
    path_2007 = MECON_DIR / "aif_spn_sector_p_blico_nacional_valores_mensuales_2007_2014.csv"
    path_apn  = MECON_DIR / "aif_apn_administraci_n_p_blica_nacional_valores_mensuales_2014_2026.csv"

    print("Reading 1993-2006...")
    df93 = read_aif(path_1993, COLS_1993)
    print(f"  {len(df93)} rows, {df93['fecha'].min().date()} - {df93['fecha'].max().date()}")

    print("Reading 2007-2014...")
    df07 = read_aif(path_2007, COLS_2007)
    print(f"  {len(df07)} rows, {df07['fecha'].min().date()} - {df07['fecha'].max().date()}")

    print("Reading 2015-2026 (APN)...")
    df_apn = read_aif(path_apn, COLS_APN)
    # Keep only from 2015 onward (SPN covers through 2014)
    df_apn = df_apn[df_apn["fecha"].dt.year >= 2015].copy()
    print(f"  {len(df_apn)} rows, {df_apn['fecha'].min().date()} - {df_apn['fecha'].max().date()}")

    # --- Concatenate ---
    monthly = pd.concat([df93, df07, df_apn], ignore_index=True)
    monthly = monthly.sort_values("fecha").reset_index(drop=True)
    print(f"\nSerie mensual combinada: {len(monthly)} filas, "
          f"{monthly['fecha'].min().date()} - {monthly['fecha'].max().date()}")

    # --- Summary check ---
    for yr in [2007, 2010, 2014, 2015, 2018]:
        sub = monthly[monthly["fecha"].dt.year == yr]
        if sub.empty:
            continue
        sp = sub["superavit_primario"].sum()
        rf = sub["resultado_financiero"].sum()
        ing = sub["ingresos_totales"].sum()
        print(f"  {yr}: Ingresos={ing:,.0f}  SP={sp:,.0f}  RF={rf:,.0f}")

    # --- Quarterly and annual ---
    quarterly = aggregate_period(monthly, "Q")
    annual    = aggregate_period(monthly, "Y")

    print(f"\nTrimestral: {len(quarterly)} filas")
    print(f"Anual:      {len(annual)} filas")

    # Write Excel (transpose: variables as rows, time as columns)
    dfs = {
        "Mensual":     monthly,
        "Trimestral":  quarterly,
        "Anual":       annual,
    }
    write_excel(dfs, OUTPUT_PATH)


if __name__ == "__main__":
    main()
