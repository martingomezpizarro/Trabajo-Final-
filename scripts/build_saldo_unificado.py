"""
build_saldo_unificado.py

Unifica saldo_deuda_nuevo.xlsx (SIGADE 2007-2018, trimestral, Adm. Central)
con boletin_mensual_31_03_2026_1.xlsx (2019-2026, mensual, SPN).

Nota de cobertura:
  - 2007-2018: Administración Central (fuente SIGADE .mdb)
  - 2019-2026: Sector Público Nacional (fuente Boletín de Deuda, MECON)

Estructura de filas: TITULOS PUBLICOS / LETRAS / PRESTAMOS (igual a A.1 SIGADE)
  La sección I (Normal) suma Mediano/Largo Plazo + Corto Plazo del boletín.
  AVALES es nueva categoría presente sólo desde 2019.

Output: data/Variables Finales/saldo_deuda_unificado.xlsx
"""

import re
import sys
import unicodedata
import warnings
from pathlib import Path
from datetime import date

import pandas as pd
import numpy as np
import openpyxl

# Force UTF-8 output to avoid cp1252 encoding errors on Windows
if sys.stdout.encoding != "utf-8":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")


def _nk(s: str) -> str:
    """Accent-strip + upper for robust label comparison.
    Also removes Unicode replacement char (�) that appears when
    openpyxl reads accented chars in cp1252-encoded files.
    """
    s = str(s).replace("�", "I")   # garbled I�I -> restore as I (cp1252 artifact)
    return "".join(
        c for c in unicodedata.normalize("NFD", s.upper())
        if unicodedata.category(c) != "Mn"
    ).strip()

warnings.filterwarnings("ignore", category=UserWarning)

SIGADE_PATH  = Path(r"c:\Users\Usuario\Desktop\MARTIN\ECONOMICS\TRABAJO FINAL\Trabajo\data\Variables Finales\saldo_deuda_nuevo.xlsx")
BOLETIN_PATH = Path(r"c:\Users\Usuario\Desktop\MARTIN\ECONOMICS\TRABAJO FINAL\Trabajo\data\raw\mecon\boletin_mensual_31_03_2026_1.xlsx")
OUTPUT_PATH  = Path(r"c:\Users\Usuario\Desktop\MARTIN\ECONOMICS\TRABAJO FINAL\Trabajo\data\Variables Finales\saldo_deuda_unificado.xlsx")

# ============================================================
# Boletín A.1 row numbers → our key names
# ============================================================
# Each entry: our_label -> list of boletín Excel rows to sum
BOLETIN_ROWS = {
    # Totales
    "total_bruto":       [10],   # A- DEUDA BRUTA (I+II+III)
    # I - Normal (ML + CP combined)
    "I_normal":          [15],
    "I_titulos":         [19, 136],    # TITULOS PUBLICOS ML + CP
    "I_letras":          [83, 138],    # LETRAS DEL TESORO ML + CP
    "I_prestamos":       None,         # computed as I_normal - I_titulos - I_letras
    "I_prest_garantiz":  [97],
    "I_org_int":         [99],
    "I_BCIE":            [100],
    "I_BIRF":            [101],
    "I_BID":             [102],
    "I_FONPLATA":        [103],
    "I_FIDA":            [104],
    "I_CAF":             [105],
    "I_OFID":            [106],
    "I_FMI":             [107],
    "I_BEI":             [108],
    "I_org_of":          [110],
    "I_CLUB_PARIS":      [111],
    "I_BILATERALES":     [112],        # "OTROS BILATERALES"
    "I_banca":           [114, 129],   # BANCA COMERCIAL ML + CP
    "I_pagares":         [116, 131],   # PAGARES DEL TESORO ML + CP
    "I_avales":          [121, 134],   # AVALES ML + CP (sólo desde 2019)
    "I_anticipos_bcra":  [123, 127],   # Extraordinarios (ML) + Ordinarios (CP)
    # II - Diferido
    "II_diferido":       [155],
    # III - Elegible
    "III_elegible":      [160],
}

# ============================================================
# Row definitions for the unified output
# (label, sigade_index_label, boletin_key)
# sigade_index_label: the exact string in SIGADE df.index, or None (computed)
# boletin_key: key in BOLETIN_ROWS, or None (use sigade only / computed)
# ============================================================
ROW_DEFS = [
    # Grand total
    ("DEUDA BRUTA TOTAL (I + II + III)",          "DEUDA BRUTA TOTAL (I + II + III)",  "total_bruto",    0),
    # I - Normal
    ("I.   DEUDA EN SITUACION DE PAGO NORMAL",    "I.   DEUDA EN SITUACION DE PAGO NORMAL", "I_normal", 0),
    ("  TITULOS PUBLICOS",                         "  TITULOS PUBLICOS",                "I_titulos",      2),
    ("    BONOS",                                  "    BONOS",                         "I_titulos",      4),  # TITULOS=BONOS in SIGADE
    ("  LETRAS",                                   "  LETRAS",                          "I_letras",       2),
    ("    LETRAS DEL TESORO",                      "    LETRAS DEL TESORO",             "I_letras",       4),
    ("    LETRAS EN GARANTÍA",                     "    LETRAS EN GARANT�A",       None,             4),  # SIGADE only (no separate line in boletín)
    ("  PRESTAMOS",                                "  PRESTAMOS",                       "I_prestamos",    2),
    ("    PRESTAMOS GARANTIZADOS",                 "    PRESTAMOS GARANTIZADOS",        "I_prest_garantiz",4),
    ("    ORGANISMOS INTERNACIONALES",             "    ORGANISMOS INTERNACIONALES",    "I_org_int",      4),
    ("      BCIE",                                 "      BCIE",                        "I_BCIE",         6),
    ("      BEI",                                  "      BEI",                         "I_BEI",          6),
    ("      BID",                                  "      BID",                         "I_BID",          6),
    ("      BIRF",                                 "      BIRF",                        "I_BIRF",         6),
    ("      CAF",                                  "      CAF",                         "I_CAF",          6),
    ("      FIDA",                                 "      FIDA",                        "I_FIDA",         6),
    ("      FMI",                                  "      FMI",                         "I_FMI",          6),
    ("      FONPLATA",                             "      FONPLATA",                    "I_FONPLATA",     6),
    ("      OFID",                                 "      OFID",                        "I_OFID",         6),
    ("    ORGANISMOS OFICIALES",                   "    ORGANISMOS OFICIALES",          "I_org_of",       4),
    ("      BILATERALES",                          "      BILATERALES",                 "I_BILATERALES",  6),
    ("      CLUB DE PARIS",                        "      CLUB DE PARIS",               "I_CLUB_PARIS",   6),
    ("    BANCA COMERCIAL",                        "    BANCA COMERCIAL",               "I_banca",        4),
    ("    PAGARES DEL TESORO",                     "    PAGARES DEL TESORO",            "I_pagares",      4),
    ("    ANTICIPOS BCRA",                         "    ANTICIPOS BCRA",                "I_anticipos_bcra",4),
    ("    AVALES (1)",                             None,                                "I_avales",       4),  # Boletín only
    # II - Diferido
    ("II.  DEUDA EN SITUACION DE PAGO DIFERIDO",  "II.  DEUDA EN SITUACION DE PAGO DIFERIDO", "II_diferido", 0),
    # III - Elegible
    ("III. DEUDA ELEGIBLE PENDIENTE DE REESTRUCTURACION",
                                                   "III. DEUDA ELEGIBLE PENDIENTE DE REESTRUCTURACION",
                                                                                        "III_elegible",   0),
]

# ============================================================

def col_to_date(col: str) -> date:
    mm, yyyy = str(col).strip().split("/")
    return date(int(yyyy), int(mm), 1)


def read_sigade(path: Path) -> pd.DataFrame:
    """Read A.1 from saldo_deuda_nuevo.xlsx. Returns df with index=labels, cols=periods."""
    df = pd.read_excel(path, sheet_name="A.1", skiprows=7, index_col=0)
    df.columns = [str(c).strip() for c in df.columns]
    date_cols = [c for c in df.columns if re.match(r"^\d{2}/\d{4}$", c)]
    df = df[date_cols]
    df.index = df.index.astype(str)
    df = df.apply(pd.to_numeric, errors="coerce")
    return df


def read_boletin(path: Path) -> dict[str, pd.Series]:
    """
    Read A.1 from boletín. Returns dict: boletin_key -> Series (date_col -> value).
    Only includes date columns (MM/YYYY), not % columns.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["A.1"]

    # Find date columns in row 9
    date_pattern = re.compile(r"^\d{2}/\d{4}$")
    col_map = {}   # col_index -> period_label
    for c in range(1, ws.max_column + 1):
        v = ws.cell(9, c).value
        if v is None:
            continue
        if hasattr(v, "strftime"):
            lbl = v.strftime("%m/%Y")
        else:
            lbl = str(v).strip()
        if date_pattern.match(lbl):
            col_map[c] = lbl

    periods = list(col_map.values())
    print(f"  Boletín periods: {periods[0]} ... {periods[-1]}  ({len(periods)})")

    # Read each row set we need
    result = {}
    for key, rows in BOLETIN_ROWS.items():
        if rows is None:
            continue
        series_data = {p: 0.0 for p in periods}
        for row_idx in rows:
            for c, period in col_map.items():
                v = ws.cell(row_idx, c).value
                if v is not None:
                    try:
                        series_data[period] += float(v)
                    except (ValueError, TypeError):
                        pass
        result[key] = pd.Series(series_data)

    # Compute I_prestamos = I_normal - I_titulos - I_letras
    result["I_prestamos"] = result["I_normal"] - result["I_titulos"] - result["I_letras"]

    # Sanity check: find last period with actual data
    total = result["total_bruto"]
    nonempty = [p for p in periods if total.get(p, 0) > 0]
    last = nonempty[-1] if nonempty else "N/A"
    print(f"  Boletín 01/2019 total: {total.get('01/2019', 0):,.1f} M USD")
    print(f"  Boletín last non-zero period: {last}  ({total.get(last, 0):,.1f} M USD)")

    return result, periods


def build_unified(sigade_df: pd.DataFrame,
                  boletin_data: dict[str, pd.Series],
                  boletin_periods: list[str]) -> pd.DataFrame:
    """
    SIGADE index has duplicate labels (e.g. '    BONOS' appears in I, II, III).
    We use sequential positional scanning: each ROW_DEF is matched to the first
    SIGADE row at or after the previous match, so ordering is preserved correctly.
    """
    sigade_cols  = list(sigade_df.columns)
    sigade_index = list(sigade_df.index)      # list of labels, duplicates allowed
    all_cols     = sigade_cols + boletin_periods

    labels = [r[0] for r in ROW_DEFS]
    data   = {col: [np.nan] * len(labels) for col in all_cols}

    sigade_search_from = 0   # advance sequentially to avoid wrong-section matches

    for i, (label, sigade_lbl, bol_key, _indent) in enumerate(ROW_DEFS):

        # --- SIGADE columns: positional scan ---
        if sigade_lbl is not None:
            matched_pos = None
            target_nk = _nk(sigade_lbl)
            for pos in range(sigade_search_from, len(sigade_index)):
                if _nk(sigade_index[pos]) == target_nk:
                    matched_pos = pos
                    break

            if matched_pos is not None:
                sigade_search_from = matched_pos + 1
                row_series = sigade_df.iloc[matched_pos]
                for col in sigade_cols:
                    v = row_series[col]
                    data[col][i] = float(v) if pd.notna(v) else np.nan
            else:
                print(f"  [WARN] SIGADE row not found: {ascii(sigade_lbl)}")

        # --- Boletín columns ---
        if bol_key is not None and bol_key in boletin_data:
            series = boletin_data[bol_key]
            for col in boletin_periods:
                v = series.get(col, np.nan)
                if v is not None and not (isinstance(v, float) and np.isnan(v)):
                    data[col][i] = float(v)

    # Build DataFrame
    result = pd.DataFrame(data, index=range(len(labels)), columns=all_cols, dtype=float)
    result.insert(0, "", labels)
    result = result.set_index("")
    result.index.name = ""
    return result


def _write_sheet(ws, df: pd.DataFrame, title: str, subtitle: str, note: str) -> None:
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    BOLD_AT_LEADING = {0, 2}       # leading-space counts that are bold
    BOLD_SUB_HEADERS = {"ORGANISMOS INTERNACIONALES", "ORGANISMOS OFICIALES"}

    ws.cell(1, 1, "MINISTERIO DE ECONOMIA")
    ws.cell(2, 1, "SECRETARIA DE FINANZAS")
    ws.cell(4, 1, title)
    ws.cell(5, 1, subtitle)
    ws.cell(6, 1, note)
    ws.cell(7, 1, "Nota: 2007-2018 = Administración Central (SIGADE) | 2019-2026 = Sector Público Nacional (Boletín MECON)")

    HDR_ROW       = 9
    DATA_START    = 10
    LABEL_COL     = 1

    cols = list(df.columns)

    # Color bands: sigade cols one shade, boletin another
    sigade_fill  = PatternFill("solid", fgColor="DDEEFF")
    boletin_fill = PatternFill("solid", fgColor="E2EFDA")

    # Write period headers
    sigade_end_col = None
    for j, col in enumerate(cols):
        c = j + 2
        cell = ws.cell(HDR_ROW, c, col)
        cell.font = Font(bold=True, size=8)
        cell.alignment = Alignment(horizontal="center")
        try:
            dt = col_to_date(col)
            cell.fill = sigade_fill if dt.year <= 2018 else boletin_fill
        except Exception:
            pass
        if col == "12/2018":
            sigade_end_col = c

    # Vertical separator after last SIGADE column
    if sigade_end_col:
        for r in range(HDR_ROW, DATA_START + len(df)):
            ws.cell(r, sigade_end_col).border = openpyxl.styles.Border(
                right=openpyxl.styles.Side(style="medium")
            )

    # Write data rows
    for i, (label, row_data) in enumerate(df.iterrows()):
        r = DATA_START + i
        lbl_str     = str(label)
        lbl_stripped = lbl_str.strip()
        leading_sp  = len(lbl_str) - len(lbl_str.lstrip(" "))

        ws.cell(r, LABEL_COL, lbl_str)

        is_bold = (
            leading_sp in BOLD_AT_LEADING or
            "TOTAL" in lbl_stripped or
            "DEUDA BRUTA" in lbl_stripped or
            lbl_stripped.startswith("I.") or
            lbl_stripped.startswith("II.") or
            lbl_stripped.startswith("III.") or
            (leading_sp == 4 and lbl_stripped in BOLD_SUB_HEADERS)
        )
        ws.cell(r, LABEL_COL).font = Font(bold=is_bold)

        for j, col in enumerate(cols):
            val = row_data[col]
            if pd.notna(val) and val != 0:
                cell = ws.cell(r, 2 + j, round(float(val), 1))
                cell.number_format = "#,##0.0"
                try:
                    dt = col_to_date(col)
                    cell.fill = sigade_fill if dt.year <= 2018 else boletin_fill
                except Exception:
                    pass

    # Column widths
    ws.column_dimensions[get_column_letter(LABEL_COL)].width = 48
    for j in range(len(cols)):
        ws.column_dimensions[get_column_letter(2 + j)].width = 10

    ws.freeze_panes = ws.cell(DATA_START, 2)


def main():
    print("Reading SIGADE data...")
    sigade = read_sigade(SIGADE_PATH)
    sigade_cols = list(sigade.columns)
    print(f"  SIGADE: {len(sigade)} rows, {sigade_cols[0]} - {sigade_cols[-1]} ({len(sigade_cols)} periods)")

    print("Reading boletín data...")
    boletin_data, boletin_periods = read_boletin(BOLETIN_PATH)

    print("Building unified DataFrame...")
    unified = build_unified(sigade, boletin_data, boletin_periods)
    print(f"  Unified: {unified.shape[0]} rows × {unified.shape[1]} columns")

    # Spot-check
    total_row = "DEUDA BRUTA TOTAL (I + II + III)"
    if total_row in unified.index:
        s = unified.loc[total_row]
        print(f"  Total 12/2018 (SIGADE): {s.get('12/2018', np.nan):,.1f}")
        print(f"  Total 01/2019 (Boletín): {s.get('01/2019', np.nan):,.1f}")
        print(f"  Total 03/2026 (Boletín): {s.get('03/2026', np.nan):,.1f}")

    # Write Excel
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    from openpyxl import Workbook
    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("A.1")
    _write_sheet(ws, unified,
                 title="DEUDA BRUTA POR INSTRUMENTO Y SITUACION DE PAGO",
                 subtitle="SERIE UNIFICADA 2007-2026",
                 note="Datos en millones de U$S")

    wb.save(OUTPUT_PATH)
    print(f"\nSaved -> {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
