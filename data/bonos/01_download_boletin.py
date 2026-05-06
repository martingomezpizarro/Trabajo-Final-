"""
01_download_boletin.py — Descarga el boletín mensual de deuda de Finanzas
y explora su estructura para encontrar el perfil de vencimientos.

Ejecutar: python data/bonos/01_download_boletin.py
"""
import requests
import os

BONOS_DIR = os.path.dirname(os.path.abspath(__file__))
RAW_DIR = os.path.join(BONOS_DIR, "raw")
os.makedirs(RAW_DIR, exist_ok=True)

# Boletín mensual más reciente (marzo 2026)
URL_BOLETIN = "https://www.argentina.gob.ar/sites/default/files/boletin_mensual_31_03_2026_1.xlsx"

dest = os.path.join(RAW_DIR, "boletin_mensual_mar2026.xlsx")

print(f"Descargando boletín mensual de deuda...")
print(f"URL: {URL_BOLETIN}")
try:
    r = requests.get(URL_BOLETIN, timeout=60)
    r.raise_for_status()
    with open(dest, "wb") as f:
        f.write(r.content)
    print(f"✅ Descargado: {dest} ({len(r.content)} bytes)")
except Exception as e:
    print(f"❌ Error: {e}")
    print("   Intentá descargar manualmente desde:")
    print(f"   {URL_BOLETIN}")
    print(f"   Y guardalo en: {dest}")

# Explorar estructura si se descargó
if os.path.exists(dest):
    try:
        import openpyxl
        wb = openpyxl.load_workbook(dest, read_only=True, data_only=True)
        print(f"\n📊 Hojas encontradas ({len(wb.sheetnames)}):")
        for i, name in enumerate(wb.sheetnames):
            print(f"   {i}: '{name}'")
        
        # Leer primeras filas de cada hoja para entender estructura
        for name in wb.sheetnames:
            ws = wb[name]
            print(f"\n=== Hoja: '{name}' ===")
            row_count = 0
            for row in ws.iter_rows(max_row=5, values_only=True):
                vals = [v for v in row[:8] if v is not None]
                if vals:
                    print(f"   Fila {row_count}: {vals}")
                row_count += 1
        wb.close()
    except Exception as e:
        print(f"Error explorando: {e}")
