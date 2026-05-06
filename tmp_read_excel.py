"""Script temporal para explorar el archivo series (3).xlsm"""
import openpyxl

print("Cargando archivo series (3).xlsm ...")
wb = openpyxl.load_workbook('series (3).xlsm', read_only=True, data_only=True)
print(f"Hojas: {wb.sheetnames}")
print()

# Buscar hojas que tengan que ver con depositos o dolares
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n=== Hoja: '{sheet_name}' ===")
    # Leer las primeras 15 filas para entender la estructura
    row_count = 0
    for row in ws.iter_rows(max_row=15, values_only=False):
        values = [cell.value for cell in row]
        # Filtrar filas vacias
        if any(v is not None for v in values):
            print(f"  Fila {row[0].row}: {values[:10]}")
        row_count += 1

wb.close()
print("\nDone!")
